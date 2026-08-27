#!/usr/bin/env python3
"""Find duplicate facility rows in a Brazil-style list via registry ids.

Motivation (list 9614, Trase Brazil slaughterhouses, 2026-08-26): when every
address in a list is "CITY, STATE", similarity-based duplicate detection
inverts into noise — all 40 of ContriBot's severity-100 pairs on that list
were false positives, while 22 real duplicates were only findable through
the registry identifiers embedded in ``internal_id``: CNPJ (company tax id)
and SIE/SIF/SIM/CONSORCIO sanitary-inspection registrations, which stay with
the physical operation even when the operating company changes.

This script reproduces that method. It reads the list file (CSV, or XLSX if
openpyxl is installed), parses registry ids out of ``internal_id``, and
reports candidate pairs in three tiers:

  TIER1_CONFIRMED     same CNPJ (and usually same registration) at the same
                      site with recognizably the same name -> remove one row
  TIER2_PROBABLE      same inspection registration (scoped by state), but a
                      DIFFERENT CNPJ -> likely predecessor/successor
                      re-registration of one site; needs a human ruling
  TIER3_INCONSISTENT  the registry contradicts itself (same CNPJ far apart,
                      or same CNPJ + same spot under an unrelated name)
                      -> ask the contributor
  REVIEW              shares ids but falls between thresholds -> eyeball it

Usage:
    python3 find_registry_duplicates.py INPUT.csv [--out report.csv]
    python3 find_registry_duplicates.py INPUT.xlsx --sheet Sheet1
    python3 find_registry_duplicates.py --selftest

The input needs an ``internal_id`` column; ``name``, ``address``, ``status``
and coordinate columns are used when present (header names are matched
case-insensitively, with common variants). Row numbers in the report follow
the ContriBot convention: header = row 1, first data row = 2.

No dependencies beyond the standard library (openpyxl only for .xlsx input).
"""

import argparse
import csv
import itertools
import math
import re
import sys
from collections import defaultdict

REG_TYPES = {'SIE', 'SIF', 'SIM', 'CONSORCIO', 'CIF', 'SISBI'}

NAME_COLS = {'name', 'facility_name'}
ADDRESS_COLS = {'address', 'facility_address'}
STATUS_COLS = {'status'}
INTERNAL_ID_COLS = {'internal_id', 'internal id'}
LAT_COLS = {'lat', 'latitude'}
LNG_COLS = {'lng', 'lon', 'long', 'longitude'}

# Classification thresholds (from the 9614 manual review):
SAME_SITE_KM = 1.0        # <= this apart counts as "same site"
FAR_APART_KM = 5.0        # >= this apart with same CNPJ is an inconsistency
NAME_SAME = 0.5           # token-overlap >= this counts as "same name"
NAME_DIFFERENT = 0.3      # token-overlap < this counts as "different name"
MAX_REG_GROUP = 8         # registration groups larger than this are treated
                          # as unparseable/malformed and skipped with a note


def norm_header(h):
    return re.sub(r'\s+', ' ', (h or '').strip().lower())


def pick_column(headers, wanted):
    for i, h in enumerate(headers):
        if norm_header(h) in wanted:
            return i
    return None


def normalized_name_tokens(name):
    """Tokens of the company part of the name.

    Trase-style names embed type and city ("COMPANY - MEAT PROCESSING UNIT -
    CITY"); only the first segment identifies the company.
    """
    company = re.split(r'\s+-\s+', name or '')[0]
    tokens = re.findall(r'[A-Z0-9]+', company.upper())
    stop = {'DE', 'DA', 'DO', 'DOS', 'DAS', 'E', 'AND', 'LTDA', 'CIA',
            'EIRELI', 'EIRELLI', 'SA', 'S', 'A', 'ME', 'EPP', 'IND',
            'COM', 'INDUSTRIA', 'COMERCIO'}
    return {t for t in tokens if t not in stop}


def name_similarity(a, b):
    ta, tb = normalized_name_tokens(a), normalized_name_tokens(b)
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def haversine_km(lat1, lng1, lat2, lng2):
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lng2 - lng1)
    a = (math.sin(dp / 2) ** 2
         + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2)
    return 2 * r * math.asin(math.sqrt(a))


def parse_internal_id(cell):
    """Extract registry identifiers from an internal_id cell.

    Observed grammar per id: TYPE_CNPJ_REGNUM[_SEQ][_COMMODITY], where CNPJ
    may be empty (malformed rows) or an 11-digit CPF for sole proprietors.
    A cell may hold several ids separated by comma/semicolon/pipe/space.

    Returns (cnpjs, registrations): a set of 14-digit CNPJs (11-digit CPFs
    are collected too — they identify a person, which is just as good), and
    a set of (regtype, regnum) tuples.
    """
    cnpjs, regs = set(), set()
    for chunk in re.split(r'[,;|\s]+', (cell or '').strip()):
        if not chunk:
            continue
        parts = chunk.split('_')
        regtype = parts[0].upper() if parts and parts[0] else None
        numbers = [p for p in parts[1:] if p.isdigit()]
        for n in numbers:
            if len(n) in (11, 14):
                cnpjs.add(n)
        if regtype in REG_TYPES:
            regnum = next(
                (n for n in numbers if len(n) not in (11, 14)), None
            )
            if regnum and int(regnum) > 0:
                regs.add((regtype, str(int(regnum))))
    return cnpjs, regs


def state_of(address):
    parts = [p.strip() for p in (address or '').split(',') if p.strip()]
    return parts[-1].upper() if parts else ''


def city_of(address):
    parts = [p.strip() for p in (address or '').split(',') if p.strip()]
    return parts[0].upper() if parts else ''


class Row:
    __slots__ = ('rn', 'name', 'address', 'status', 'lat', 'lng',
                 'cnpjs', 'regs')

    def __init__(self, rn, name, address, status, lat, lng, internal_id):
        self.rn = rn
        self.name = name or ''
        self.address = address or ''
        self.status = (status or '').strip().upper() or '(blank)'
        self.lat, self.lng = lat, lng
        self.cnpjs, self.regs = parse_internal_id(internal_id)


def load_rows(path, sheet=None):
    if path.lower().endswith(('.xlsx', '.xlsm')):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit('Reading .xlsx needs openpyxl (pip install openpyxl) '
                     '— or export the sheet as CSV.')
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        raw = [[('' if c is None else str(c)) for c in r]
               for r in ws.iter_rows(values_only=True)]
    else:
        with open(path, newline='', encoding='utf-8-sig') as f:
            raw = list(csv.reader(f))
    if not raw:
        sys.exit('Empty input.')

    headers = raw[0]
    idx = {
        'internal_id': pick_column(headers, INTERNAL_ID_COLS),
        'name': pick_column(headers, NAME_COLS),
        'address': pick_column(headers, ADDRESS_COLS),
        'status': pick_column(headers, STATUS_COLS),
        'lat': pick_column(headers, LAT_COLS),
        'lng': pick_column(headers, LNG_COLS),
    }
    if idx['internal_id'] is None:
        sys.exit('No internal_id column found — headers were: %s' % headers)

    def cell(r, key):
        i = idx[key]
        return r[i] if i is not None and i < len(r) else ''

    rows = []
    for n, r in enumerate(raw[1:], start=2):  # rn convention: header = 1
        lat = lng = None
        try:
            lat, lng = float(cell(r, 'lat')), float(cell(r, 'lng'))
        except (TypeError, ValueError):
            pass
        rows.append(Row(n, cell(r, 'name'), cell(r, 'address'),
                        cell(r, 'status'), lat, lng, cell(r, 'internal_id')))
    return rows


def classify(a, b, shared_cnpj, shared_reg):
    sim = name_similarity(a.name, b.name)
    dist = None
    if None not in (a.lat, a.lng, b.lat, b.lng):
        dist = haversine_km(a.lat, a.lng, b.lat, b.lng)
    same_site = (dist is not None and dist <= SAME_SITE_KM) or (
        dist is None and city_of(a.address) == city_of(b.address))

    if shared_cnpj:
        if dist is not None and dist >= FAR_APART_KM:
            tier = 'TIER3_INCONSISTENT'
            why = 'same CNPJ but sites %.1f km apart' % dist
        elif same_site and sim >= NAME_SAME:
            tier = 'TIER1_CONFIRMED'
            why = 'same CNPJ, same site, same company name'
        elif same_site and sim < NAME_DIFFERENT:
            tier = 'TIER3_INCONSISTENT'
            why = 'same CNPJ + same site under an unrelated name'
        else:
            tier = 'REVIEW'
            why = 'same CNPJ; name/site evidence between thresholds'
    else:  # shared registration only
        tier = 'TIER2_PROBABLE'
        why = ('same %s registration %s, different CNPJ — likely '
               'predecessor/successor re-registration' % shared_reg)
        if sim >= NAME_SAME and same_site:
            tier = 'TIER1_CONFIRMED'
            why = ('same %s registration %s at the same site with the same '
                   'company name (CNPJ differs/missing on one side)'
                   % shared_reg)

    return tier, sim, dist, why


def find_pairs(rows):
    by_cnpj = defaultdict(list)
    by_reg = defaultdict(list)
    for row in rows:
        for c in row.cnpjs:
            by_cnpj[c].append(row)
        st = state_of(row.address)
        for reg in row.regs:
            by_reg[(reg, st)].append(row)

    seen, results, skipped_groups = set(), [], []
    def emit(a, b, shared_cnpj, shared_reg):
        key = tuple(sorted((a.rn, b.rn)))
        if key in seen or a.rn == b.rn:
            return
        seen.add(key)
        tier, sim, dist, why = classify(a, b, shared_cnpj, shared_reg)
        results.append((tier, a, b, sim, dist, why))

    for c, group in by_cnpj.items():
        for a, b in itertools.combinations(group, 2):
            emit(a, b, c, None)
    for (reg, st), group in by_reg.items():
        if len(group) > MAX_REG_GROUP:
            skipped_groups.append((reg, st, len(group)))
            continue
        for a, b in itertools.combinations(group, 2):
            if a.cnpjs & b.cnpjs:
                continue  # already emitted via CNPJ
            emit(a, b, None, reg)
    return results, skipped_groups


TIER_ORDER = ['TIER1_CONFIRMED', 'TIER2_PROBABLE', 'TIER3_INCONSISTENT',
              'REVIEW']


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument('input', nargs='?', help='CSV or XLSX list file')
    ap.add_argument('--sheet', help='worksheet name for XLSX input')
    ap.add_argument('--out', help='write pair report as CSV here')
    ap.add_argument('--selftest', action='store_true',
                    help='run built-in synthetic checks and exit')
    args = ap.parse_args(argv)

    if args.selftest:
        return selftest()
    if not args.input:
        ap.error('input file required (or --selftest)')

    rows = load_rows(args.input, args.sheet)
    results, skipped = find_pairs(rows)
    results.sort(key=lambda r: (TIER_ORDER.index(r[0]), r[1].rn))

    counts = defaultdict(int)
    for tier, *_ in results:
        counts[tier] += 1

    print('%d data rows; %d rows carry a parseable CNPJ/CPF; %d carry a '
          'registration id' % (
              len(rows),
              sum(1 for r in rows if r.cnpjs),
              sum(1 for r in rows if r.regs)))
    print('Candidate duplicate pairs: '
          + ', '.join('%s: %d' % (t, counts[t])
                      for t in TIER_ORDER if counts[t]))
    for reg, st, n in skipped:
        print('NOTE: skipped registration group %s/%s in %s (%d rows — '
              'likely malformed ids)' % (reg[0], reg[1], st or '?', n))
    print('CAUTION: name/address similarity flags (ContriBot "identical/'
          'very similar rows") are unreliable on city-level-address lists — '
          'verify them against these registry-id pairs, not the other way '
          'around.')

    out_rows = [(t, a.rn, b.rn, a.name, b.name, a.status, b.status,
                 '%.2f' % s, '' if d is None else '%.2f' % d, why)
                for t, a, b, s, d, why in results]
    header = ['tier', 'row_a', 'row_b', 'name_a', 'name_b', 'status_a',
              'status_b', 'name_similarity', 'distance_km', 'evidence']
    if args.out:
        with open(args.out, 'w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(header)
            w.writerows(out_rows)
        print('Report written to %s (%d pairs)' % (args.out, len(out_rows)))
    else:
        w = csv.writer(sys.stdout)
        w.writerow(header)
        w.writerows(out_rows)
    return 0


def selftest():
    def row(rn, name, addr, status, lat, lng, iid):
        return Row(rn, name, addr, status, lat, lng, iid)

    rows = [
        # T1: same CNPJ, same site, same company, commodity re-listing
        row(2, 'ACME CARNES - MEAT PROCESSING UNIT - CURITIBA',
            'CURITIBA, PARANA', 'ACTIVE', -25.4, -49.2,
            'SIE_11122233344455_101_1_PORK'),
        row(3, 'ACME CARNES - CANNING FACTORY - CURITIBA',
            'CURITIBA, PARANA', '', -25.4, -49.2,
            'SIE_11122233344455_101_2_CANNEDGOODS'),
        # T2: same registration, different CNPJ, ACTIVE/INACTIVE
        row(4, 'OLDCO ALIMENTOS - MEAT PROCESSING UNIT - MARINGA',
            'MARINGA, PARANA', 'INACTIVE', -23.42, -51.93,
            'SIE_99887766554433_202_1_BEEF'),
        row(5, 'NEWCO ALIMENTOS - MEAT PROCESSING UNIT - MARINGA',
            'MARINGA, PARANA', 'ACTIVE', -23.421, -51.931,
            'SIE_55667788990011_202_1_BEEF'),
        # T3: same CNPJ, 100+ km apart
        row(6, 'LONGE LATICINIOS - MILK - GASPAR', 'GASPAR, SANTA CATARINA',
            'ACTIVE', -26.93, -48.95, 'SIE_12312312312312_303_1_DAIRY'),
        row(7, 'LONGE LATICINIOS - MILK - TAIO', 'TAIO, SANTA CATARINA',
            '', -27.12, -50.0, 'SIE_12312312312312_304_1_DAIRY'),
        # T3: same CNPJ + site, unrelated names
        row(8, 'WOLLEN CHARCUTARIA - MEAT - NITEROI',
            'NITEROI, RIO DE JANEIRO', 'ACTIVE', -22.88, -43.10,
            'SIE_45645645645645_405_1_PORK'),
        row(9, 'REM EMBUTIDOS - CANNING - NITEROI',
            'NITEROI, RIO DE JANEIRO', '', -22.88, -43.10,
            'SIE_45645645645645_405_2_CANNEDGOODS'),
        # not a pair: same regnum, DIFFERENT state
        row(10, 'SULINA CARNES - MEAT - PORTO ALEGRE',
            'PORTO ALEGRE, RIO GRANDE DO SUL', 'ACTIVE', -30.03, -51.23,
            'SIE_77777777777777_202_1_BEEF'),
        # malformed id (empty CNPJ) still yields a registration
        row(11, 'AKROPOLIS - SLAUGHTERHOUSE - SAO GONCALO',
            'SAO GONCALO, RIO DE JANEIRO', 'ACTIVE', None, None,
            'SIE__505_1_PORK'),
        row(12, 'AGROPECUARIA AKROPOLIS - SLAUGHTERHOUSE - SAO GONCALO',
            'SAO GONCALO, RIO DE JANEIRO', '', None, None,
            'SIE_88888888888888_505_1_PORK'),
    ]
    results, skipped = find_pairs(rows)
    got = {tuple(sorted((a.rn, b.rn))): t for t, a, b, *_ in results}
    expected = {
        (2, 3): 'TIER1_CONFIRMED',
        (4, 5): 'TIER2_PROBABLE',
        (6, 7): 'TIER3_INCONSISTENT',
        (8, 9): 'TIER3_INCONSISTENT',
        (11, 12): 'TIER1_CONFIRMED',  # same reg+state+site+name, one CNPJ
    }
    ok = True
    for pair, tier in expected.items():
        if got.get(pair) != tier:
            print('FAIL %s: expected %s, got %s'
                  % (pair, tier, got.get(pair)))
            ok = False
    unexpected = set(got) - set(expected)
    if unexpected:
        print('FAIL unexpected pairs: %s' % sorted(unexpected))
        ok = False
    if skipped:
        print('FAIL unexpected skipped groups: %s' % skipped)
        ok = False
    print('selftest: %s (%d pairs)' % ('OK' if ok else 'FAILED', len(got)))
    return 0 if ok else 1


if __name__ == '__main__':
    sys.exit(main())
