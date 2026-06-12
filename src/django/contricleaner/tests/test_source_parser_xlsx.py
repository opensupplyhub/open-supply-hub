import os
from unittest.mock import MagicMock, patch

from defusedxml.common import EntitiesForbidden
from django.test import TestCase
from django.core.files.uploadedfile import (
    SimpleUploadedFile
)
from django.core.files.base import File
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

from contricleaner.lib.parsers.source_parser_xlsx import SourceParserXLSX
from contricleaner.lib.parsers.abstractions.source_parser import (
    SourceParser
)
from contricleaner.lib.parsers.abstractions.file_parser import (
    FileParser
)
from contricleaner.lib.exceptions.parsing_error import ParsingError
from contricleaner.lib.contri_cleaner import ContriCleaner
from contricleaner.lib.dto.list_dto import ListDTO
from contricleaner.lib.dto.row_dto import RowDTO
from contricleaner.tests.sector_cache_mock import SectorCacheMock
from contricleaner.tests.os_id_lookup_mock import OSIDLookupMock


class SourceParserXLSXTest(TestCase):
    def test_class_inherits_required_classes(self):
        temp_uploaded_file_stub = MagicMock(spec=File)
        source_parser_xlsx = SourceParserXLSX(temp_uploaded_file_stub)
        self.assertIsInstance(source_parser_xlsx, SourceParser)
        self.assertIsInstance(source_parser_xlsx, FileParser)

    def test_valid_xlsx_file_parsing(self):
        sheet_rows = (
            ('country', 'name', 'address', 'sector_product_type',
             'facility_type_processing_type', 'number_of_workers',
             'parent_company', 'percentage_of_male_workers'),
            ('United States', 'Fashion Plus',
             '123 Avenue Street,,,, Cityville,', 'Apparel|Jeans',
             'Embellishment', 1002, '', 2.00),
            ('Canada', 'N/A', '456 Main Road,,,, Townsville,',
             'n/a', 'Embossing', 1005, '', 3.50),
            ('Canada', 'Style Haven', '456 Fashion Road, Trendy Town',
             'Apparel|Jacket', '', 500, 'Style Super  ', 1),
            ('Italy', 'Chic Boutique', '789 Glamour Avenue,,,, Moda City,',
             'Apparel', 'Embossing', 3002, '', '20%')
        )

        expected_processed_rows = [
            RowDTO(
                raw_json={
                    "country": "United States",
                    "name": "Fashion Plus",
                    "address": "123 Avenue Street,,,, Cityville,",
                    "sector_product_type": "Apparel|Jeans",
                    "facility_type_processing_type": "Embellishment",
                    "number_of_workers": "1002",
                    "parent_company": "",
                    "percentage_of_male_workers": "2%",
                },
                name="Fashion Plus",
                clean_name="fashion plus",
                address="123 Avenue Street, Cityville",
                clean_address="123 avenue street cityville",
                country_code="US",
                sector=["Apparel"],
                fields={
                    "product_type": ["Jeans"],
                    "facility_type": {
                        "raw_values": "Embellishment",
                        "processed_values": ["Embellishment"],
                    },
                    "processing_type": {
                        "raw_values": "Embellishment",
                        "processed_values": ["Embellishment"],
                    },
                    "country": "United States",
                    "facility_type_processing_type": "Embellishment",
                    "number_of_workers": "1002",
                    "sector_product_type": "Apparel|Jeans",
                    "parent_company": "",
                    "percentage_of_male_workers": "2%",
                },
                errors=[],
            ),
            RowDTO(
                raw_json={
                    "country": "Canada",
                    "name": "N/A",
                    "address": "456 Main Road,,,, Townsville,",
                    "sector_product_type": "n/a",
                    "facility_type_processing_type": "Embossing",
                    "number_of_workers": "1005",
                    "parent_company": "",
                    "percentage_of_male_workers": "3.5%",
                },
                name="",
                clean_name="",
                address="456 Main Road, Townsville",
                clean_address="456 main road townsville",
                country_code="CA",
                sector=["Unspecified"],
                fields={
                    "facility_type": {
                        "raw_values": "Embossing",
                        "processed_values": ["Embossing"],
                    },
                    "processing_type": {
                        "raw_values": "Embossing",
                        "processed_values": ["Embossing"],
                    },
                    "percentage_of_male_workers": "3.5%",
                    "parent_company": "",
                    "sector_product_type": "",
                    "country": "Canada",
                    "facility_type_processing_type": "Embossing",
                    "number_of_workers": "1005",
                },
                errors=[
                    {
                        "message": (
                            "name cannot consist solely of punctuation "
                            "or whitespace."
                        ),
                        "field": "name",
                        "type": "Error"
                    }
                ],
            ),
            RowDTO(
                raw_json={
                    "country": "Canada",
                    "name": "Style Haven",
                    "address": "456 Fashion Road, Trendy Town",
                    "sector_product_type": "Apparel|Jacket",
                    "facility_type_processing_type": "",
                    "number_of_workers": "500",
                    "parent_company": "Style Super  ",
                    "percentage_of_male_workers": "100%",
                },
                name="Style Haven",
                clean_name="style haven",
                address="456 Fashion Road, Trendy Town",
                clean_address="456 fashion road trendy town",
                country_code="CA",
                sector=["Apparel"],
                fields={
                    "product_type": ["Jacket"],
                    "country": "Canada",
                    "facility_type_processing_type": "",
                    "number_of_workers": "500",
                    "sector_product_type": "Apparel|Jacket",
                    "parent_company": "Style Super",
                    "percentage_of_male_workers": "100%",
                },
                errors=[],
            ),
            RowDTO(
                raw_json={
                    "country": "Italy",
                    "name": "Chic Boutique",
                    "address": "789 Glamour Avenue,,,, Moda City,",
                    "sector_product_type": "Apparel",
                    "facility_type_processing_type": "Embossing",
                    "number_of_workers": "3002",
                    "parent_company": "",
                    "percentage_of_male_workers": "20%",
                },
                name="Chic Boutique",
                clean_name="chic boutique",
                address="789 Glamour Avenue, Moda City",
                clean_address="789 glamour avenue moda city",
                country_code="IT",
                sector=["Apparel"],
                fields={
                    "facility_type": {
                        "raw_values": "Embossing",
                        "processed_values": ["Embossing"],
                    },
                    "processing_type": {
                        "raw_values": "Embossing",
                        "processed_values": ["Embossing"],
                    },
                    "country": "Italy",
                    "facility_type_processing_type": "Embossing",
                    "number_of_workers": "3002",
                    "sector_product_type": "Apparel",
                    "parent_company": "",
                    "percentage_of_male_workers": "20%",
                },
                errors=[],
            ),
        ]
        expected_processed_list = ListDTO(rows=expected_processed_rows)

        workbook = Workbook()
        sheet = workbook.active

        for row in sheet_rows:
            sheet.append(row)

        # Set number format for percentage and format to display as percentage
        # with 2 decimal places.
        percentage_style = NamedStyle(
            name='percentage',
            number_format='0.00%'
        )

        # Apply the percentage style to the column containing percentage
        # values (index 8)
        for row in sheet.iter_rows(min_row=1, max_row=len(sheet_rows),
                                   min_col=8, max_col=8):
            for cell in row:
                cell.style = percentage_style

        workbook.save('test.xlsx')

        with open('test.xlsx', 'rb') as xlsx_file:
            file_content = xlsx_file.read()
            uploaded_file = SimpleUploadedFile('test.xlsx', file_content)

        contri_cleaner = ContriCleaner(
            uploaded_file,
            SectorCacheMock(),
            OSIDLookupMock()
        )
        processed_list = contri_cleaner.process_data()

        self.assertEqual(processed_list.rows, expected_processed_list.rows)

        os.remove('test.xlsx')

    def test_percent_formatting_detected_per_cell_not_only_from_row_2(self):
        """Regression test: percent cells must be detected per cell.

        Previously the parser only examined row 2 to decide which columns were
        percent-formatted.  If row 2 had a blank/None value in a percent column
        (e.g. the contributor left that cell empty), every other row in that
        column would be stored as a raw decimal (0.5) instead of a percentage
        (50%).  The fix checks cell.number_format on every cell individually.
        """
        # Row 2 has an empty percent cell; rows 3-4 have real decimal values
        # that must be converted to percent strings.
        sheet_rows = (
            ('country', 'name', 'address', 'percentage_of_male_workers'),
            ('United States', 'Factory A', '1 Main St', None),   # row 2: blank
            ('Canada', 'Factory B', '2 Side Ave', 0.5),          # row 3: 50%
            ('Italy', 'Factory C', '3 Via Roma', 0.75),          # row 4: 75%
        )

        workbook = Workbook()
        sheet = workbook.active
        for row in sheet_rows:
            sheet.append(row)

        percentage_style = NamedStyle(
            name='percentage',
            number_format='0.00%'
        )
        # Apply percent style to ALL data cells in column 4 (including the
        # blank row-2 cell) — this is the realistic scenario where the
        # spreadsheet template marks the column as percent-formatted.
        for row in sheet.iter_rows(min_row=1, max_row=len(sheet_rows),
                                   min_col=4, max_col=4):
            for cell in row:
                cell.style = percentage_style

        workbook.save('test_per_cell.xlsx')

        with open('test_per_cell.xlsx', 'rb') as xlsx_file:
            uploaded_file = SimpleUploadedFile(
                'test_per_cell.xlsx', xlsx_file.read()
            )

        contri_cleaner = ContriCleaner(
            uploaded_file,
            SectorCacheMock(),
            OSIDLookupMock()
        )
        processed_list = contri_cleaner.process_data()

        percent_values = [
            row.raw_json.get('percentage_of_male_workers')
            for row in processed_list.rows
        ]

        # Row 2 (blank) should remain empty; rows 3-4 must be percent strings.
        self.assertEqual(percent_values[0], '')      # blank → empty string
        self.assertEqual(percent_values[1], '50%')   # 0.5  → '50%'
        self.assertEqual(percent_values[2], '75%')   # 0.75 → '75%'

        os.remove('test_per_cell.xlsx')

    @patch('contricleaner.lib.parsers.source_parser_xlsx.load_workbook')
    def test_entities_forbidden_exception(self, mock_load_workbook):
        mock_load_workbook.side_effect = EntitiesForbidden(
            name='', value='', base='', sysid='', pubid='', notation_name=''
        )
        temp_uploaded_file_stub = MagicMock(spec=File)
        parser = SourceParserXLSX(temp_uploaded_file_stub)

        with self.assertRaisesRegex(
                ParsingError,
                ('There was an error within your file and our team needs to '
                 'take a look. Please send your file to '
                 'support@opensupplyhub.org for diagnosis.'),
                ):
            parser.get_parsed_rows()
