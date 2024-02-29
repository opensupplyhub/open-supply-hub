from typing import List
import csv
import json
from openpyxl import load_workbook
import re
from django.core.exceptions import ValidationError
from openpyxl.utils import get_column_letter

from contricleaner.lib.parsers.data_parser import DataParser


def try_parse_int_from_float(value):
    try:
        return str(int(float(value)))
    except Exception:
        return value


def get_csv_values(csv_data):
    values = []
    csvreader = csv.reader(csv_data.split('\n'), delimiter=',')
    for row in csvreader:
        values = values + row
    return values


def get_raw_json(raw_data: str, raw_header: str) -> dict:
    data_values = get_csv_values(raw_data)
    list_fields = get_csv_values(raw_header)
    fields = dict()
    for i, name in enumerate(list_fields):
        if name != '':
            value = data_values[i] if i < len(data_values) else ""
            fields[name.lower()] = try_parse_int_from_float(value)

    return fields



def replace_invalid_data(value, invalid_keywords):
    result_value = value
    for keyword in invalid_keywords:
        # Remove invalid keywords if exist
        result_value = result_value.replace(keyword, '')
    return result_value


def cleanup_data(value):
    dup_pattern = ',' + '{2,}'
    # Remove duplicates commas if exist
    result_value = re.sub(dup_pattern, ',', value)
    # Remove comma in the end of the string if exist
    result_value = result_value.rstrip(',')
    # Remove extra spaces if exist
    return result_value.strip()


def clean_row(row):
    invalid_keywords = ['N/A', 'n/a']
    replaced_value = replace_invalid_data(row, invalid_keywords)
    return cleanup_data(replaced_value)


def parse_csv_line(line):
    return list(csv.reader([line]))[0]


def get_xlsx_sheet(file, request):
    import defusedxml
    from defusedxml.common import EntitiesForbidden

    defusedxml.defuse_stdlib()

    try:
        wb = load_workbook(filename=file)
        ws = wb[wb.sheetnames[0]]

        return ws

    except EntitiesForbidden:
        # _report_error_to_rollbar(file, request)
        raise ValidationError('This file may be damaged and '
                              'cannot be processed safely')


def format_cell_value(value):
    if value is None:
        return ''
    return str(value)


def format_percent(value):
    if value is None or isinstance(value, str):
        return value
    if value <= 1.0:
        str_value = str(value * 100)
    else:
        str_value = str(value)
    if str_value[-2:] == '.0':
        str_value = str_value[0:len(str_value)-2]
    return str_value + '%'


def parse_xlsx(file, request):
    try:
        ws = get_xlsx_sheet(file, request)

        # openpyxl package is 1-indexed
        percent_col = [get_column_letter(cell.column)
                       for cell in ws[2] if '%' in cell.number_format]

        if percent_col:
            for col in percent_col:
                for cell in ws[col]:
                    if cell.row != 1:
                        cell.value = format_percent(cell.value)

        ws_rows = ws.rows
        # Useing `next` will consome the row so that iteration of the data rows
        # will skipt the header row
        first_row = next(ws_rows)
        header = ','.join(
            [format_cell_value(cell.value) for cell in first_row])

        def format_row(row):
            return '"{}"'.format(
                '","'.join([format_cell_value(cell.value) for cell in row]))

        rows = [format_row(row)
                for row in ws_rows
                if any(cell.value is not None for cell in row)]

        return header, rows
    except Exception:
        # _report_error_to_rollbar(file, request)
        raise ValidationError('Error parsing Excel (.xlsx) file')


class DataParserCSV(DataParser):
    def __init__(self, file):
        self.__file = file

    def __parse(self, file):
        rows = []

        try:
            header = file.readline().decode(encoding='utf-8-sig').rstrip()
        except UnicodeDecodeError:
            # TODO: find out how to post errors torollbar
            # _report_error_to_rollbar(file, request)
            raise ValidationError('Unsupported file encoding. Please '
                                  'submit a UTF-8 CSV.')

        for idx, line in enumerate(file):
            if idx > 0:
                try:
                    rows.append(line.decode(encoding='utf-8-sig').rstrip())
                except UnicodeDecodeError:
                    # _report_error_to_rollbar(file, request)
                    raise ValidationError('Unsupported file encoding. Please '
                                        'submit a UTF-8 CSV.')

        return header, rows

    def get_parsed_rows(self) -> List[dict]:
        header, rows = self.__parse()
        return [get_raw_json(row, header) for row in rows]
