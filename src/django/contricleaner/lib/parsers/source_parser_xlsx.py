from typing import List, Union

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from django.core.files.base import File
from django.conf import settings
from django.core.files.storage import default_storage
import os

from contricleaner.lib.parsers.abstractions.source_parser import SourceParser
from contricleaner.lib.parsers.abstractions.file_parser import FileParser
from contricleaner.lib.exceptions.parsing_error import ParsingError

import logging
import traceback
logger = logging.getLogger(__name__)


class SourceParserXLSX(SourceParser, FileParser):
    def get_parsed_rows(self) -> List[dict]:
        return self._parse(self._file)

    @staticmethod
    def _parse(file: File) -> List[dict]:
        try:
            logger.info(
                "XLSX parse: file.name=%s storage=%s storage_class=%s "
                "default_storage=%s default_storage_class=%s "
                "DEFAULT_FILE_STORAGE=%s AWS_STORAGE_BUCKET_NAME=%s "
                "DJANGO_SETTINGS_MODULE=%s DEBUG=%s",
                getattr(file, "name", "<no-name>"),
                getattr(file, "storage", None),
                getattr(getattr(file, "storage", None), "__class__", None),
                default_storage,
                default_storage.__class__,
                getattr(settings, "DEFAULT_FILE_STORAGE", None),
                getattr(settings, "AWS_STORAGE_BUCKET_NAME", None),
                os.getenv("DJANGO_SETTINGS_MODULE"),
                settings.DEBUG,
            )
            worksheet = SourceParserXLSX.__get_xlsx_sheet(file)

            # openpyxl package is 1-indexed
            percent_col = [get_column_letter(cell.column)
                           for cell in worksheet[2]
                           if '%' in cell.number_format]

            if percent_col:
                for col in percent_col:
                    for cell in worksheet[col]:
                        if cell.row != 1:
                            cell.value = SourceParserXLSX.__format_percent(
                                cell.value
                            )

            worksheet_rows = worksheet.rows
            # Using `next` extracts the header row, causing the iteration of
            # the data rows to skip the header row.
            first_row = next(worksheet_rows)
            header = [
                SourceParserXLSX.__format_cell_value(cell.value)
                for cell in first_row
            ]

            rows = [dict(zip(header, SourceParserXLSX.__tidy_row(row)))
                    for row in worksheet_rows
                    if any(cell.value is not None for cell in row)]

            return rows
        except Exception:
            logger.error(f"Error parsing XLSX file: {traceback.format_exc()}")
            raise ParsingError(
                'There was an error within your file and our team needs to '
                'take a look. Please send your file to '
                'support@opensupplyhub.org for diagnosis.')

    @staticmethod
    def __tidy_row(row: tuple) -> list:
        formatted_row = []

        for cell in row:
            formatted_cell_value = \
                SourceParserXLSX.__format_cell_value(cell.value)
            formatted_row.append(formatted_cell_value)

        return formatted_row

    @staticmethod
    def __get_xlsx_sheet(file: File) -> Worksheet:
        import defusedxml
        from defusedxml.common import EntitiesForbidden

        defusedxml.defuse_stdlib()

        try:
            workbook = load_workbook(filename=file)
            worksheet = workbook[workbook.sheetnames[0]]

            return worksheet

        except EntitiesForbidden:
            raise ParsingError('This file may be damaged and '
                               'cannot be processed safely.')

    @staticmethod
    def __format_percent(value: Union[float, int, str, None]) -> str:
        if value is None or isinstance(value, str):
            return value
        if value <= 1.0:
            str_value = str(value * 100)
        else:
            str_value = str(value)
        if str_value[-2:] == '.0':
            str_value = str_value[0:len(str_value)-2]
        return str_value + '%'

    @staticmethod
    def __format_cell_value(value) -> str:
        if value is None:
            return ''
        return str(value)
