import csv
import io
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import load_workbook

from api.isic_taxonomy.constants import ALLOWED_EXTENSIONS, REQUIRED_HEADERS
from api.isic_taxonomy.errors import (
    IsicTaxonomyError,
    IsicTaxonomyValidationError,
)

ODS_NS = {
    'table': 'urn:oasis:names:tc:opendocument:xmlns:table:1.0',
    'text': 'urn:oasis:names:tc:opendocument:xmlns:text:1.0',
}


def normalize_extension(filename: str) -> str:
    extension = Path(filename).suffix.lower()
    if extension == '.odt':
        return '.ods'
    return extension


def parse_spreadsheet(
    file_content: bytes,
    filename: str,
) -> list[dict[str, str]]:
    '''
    Parse CSV, XLSX, or ODS taxonomy spreadsheets into row dicts.

    Header keys are normalized to lowercase stripped names matching
    REQUIRED_HEADERS.
    '''
    extension = normalize_extension(filename)
    if extension not in ALLOWED_EXTENSIONS:
        allowed = ', '.join(
            sorted(ext for ext in ALLOWED_EXTENSIONS if ext != '.odt')
        )
        raise IsicTaxonomyValidationError([
            IsicTaxonomyError(
                message=(
                    f'Unsupported file type "{extension or "(none)"}". '
                    f'Accepted formats: {allowed}.'
                ),
            ),
        ])

    if extension == '.csv':
        rows = _parse_csv(file_content)
    elif extension == '.xlsx':
        rows = _parse_xlsx(file_content)
    else:
        rows = _parse_ods(file_content)

    if not rows:
        raise IsicTaxonomyValidationError([
            IsicTaxonomyError(message='The file contains no data rows.'),
        ])

    return _normalize_rows(rows)


def _parse_csv(file_content: bytes) -> list[list[str]]:
    try:
        text = file_content.decode('utf-8')
    except UnicodeDecodeError as exc:
        raise IsicTaxonomyValidationError([
            IsicTaxonomyError(
                message=(
                    'CSV must be UTF-8 encoded. Re-save the file as UTF-8 '
                    'and try again.'
                ),
            ),
        ]) from exc

    reader = csv.reader(io.StringIO(text))
    return [row for row in reader if any(cell.strip() for cell in row)]


def _parse_xlsx(file_content: bytes) -> list[list[str]]:
    try:
        workbook = load_workbook(
            filename=io.BytesIO(file_content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise IsicTaxonomyValidationError([
            IsicTaxonomyError(
                message=(
                    'Could not read the XLSX file. Ensure it is a valid, '
                    'unencrypted spreadsheet.'
                ),
            ),
        ]) from exc

    try:
        worksheet = workbook.active
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            string_row = [
                '' if value is None else str(value).strip()
                for value in row
            ]
            if any(string_row):
                rows.append(string_row)
        return rows
    finally:
        workbook.close()


def _parse_ods(file_content: bytes) -> list[list[str]]:
    try:
        with zipfile.ZipFile(io.BytesIO(file_content)) as archive:
            content = archive.read('content.xml')
    except (zipfile.BadZipFile, KeyError) as exc:
        raise IsicTaxonomyValidationError([
            IsicTaxonomyError(
                message=(
                    'Could not read the ODS file. Ensure it is a valid '
                    'OpenDocument spreadsheet.'
                ),
            ),
        ]) from exc

    root = ET.fromstring(content)
    rows = []
    for table in root.findall('.//table:table', ODS_NS):
        for row in table.findall('table:table-row', ODS_NS):
            cells = []
            for cell in row.findall('table:table-cell', ODS_NS):
                repeat = int(
                    cell.get(
                        '{urn:oasis:names:tc:opendocument:xmlns:table:1.0}'
                        'number-columns-repeated',
                        1,
                    )
                )
                value = _ods_cell_text(cell)
                cells.extend([value] * repeat)
            if any(cells):
                rows.append(cells)
    return rows


def _ods_cell_text(cell: ET.Element) -> str:
    parts = []
    for paragraph in cell.findall('.//text:p', ODS_NS):
        if paragraph.text:
            parts.append(paragraph.text)
        for child in paragraph:
            if child.tail:
                parts.append(child.tail)
    return ''.join(parts).strip()


def _normalize_rows(raw_rows: list[list[str]]) -> list[dict[str, str]]:
    header_row = [cell.strip().lower() for cell in raw_rows[0]]
    missing_headers = [
        header for header in REQUIRED_HEADERS if header not in header_row
    ]
    if missing_headers:
        raise IsicTaxonomyValidationError([
            IsicTaxonomyError(
                message=(
                    'Missing required column(s): '
                    f'{", ".join(missing_headers)}.'
                ),
            ),
        ])

    column_indexes = {
        header: header_row.index(header)
        for header in REQUIRED_HEADERS
    }
    normalized_rows = []

    for raw_row in raw_rows[1:]:
        if not any(cell.strip() for cell in raw_row):
            continue

        row = {}
        for header, index in column_indexes.items():
            if index < len(raw_row):
                row[header] = raw_row[index].strip()
            else:
                row[header] = ''
        normalized_rows.append(row)

    return normalized_rows
