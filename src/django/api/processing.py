import copy
import csv
import time
import traceback
import logging
import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from django.contrib.gis.geos import Point
from django.core.exceptions import ValidationError
from django.urls import reverse
from django.utils import timezone
from django.db import transaction

from api.constants import CsvHeaderField, ProcessingAction
from api.extended_fields import (
    create_extendedfields_for_listitem,
    update_extendedfields_for_list_item,
)
from api.geocoding import geocode_address
from api.helpers.helpers import clean, cleanup_data, replace_invalid_data
from api.matching import normalize_extended_facility_id
from api.models.facility.facility import Facility
from api.models.facility.facility_list_item import FacilityListItem
from api.models.facility.facility_list_item_temp import FacilityListItemTemp
from api.models.facility.facility_match import FacilityMatch
from api.models.facility.facility_match_temp import FacilityMatchTemp
from api.models.facility.facility_index import FacilityIndex
from api.models.facility.facility_list import FacilityList
from api.sector_product_type_parser import CsvRowSectorProductTypeParser
from contricleaner.lib.contri_cleaner import ContriCleaner
from contricleaner.lib.exceptions.parsing_error import ParsingError
from contricleaner.lib.exceptions.handler_not_set_error \
    import HandlerNotSetError
from api.sector_product_type_parser import SectorCache
from api.facility_actions.processing_facility_executor import (
    ProcessingFacilityExecutor
)
from api.facility_actions.processing_facility_list import (
    ProcessingFacilityList
)
from countries.lib.get_country_code import get_country_code
from oar.rollbar import report_error_to_rollbar

from django.core.files.uploadedfile import (
    InMemoryUploadedFile,
    TemporaryUploadedFile
)

from django.db.models.fields.files import FieldFile

logger = logging.getLogger(__name__)


def get_xlsx_sheet(file, request):
    import defusedxml
    from defusedxml.common import EntitiesForbidden

    defusedxml.defuse_stdlib()

    try:
        wb = load_workbook(filename=file)
        ws = wb[wb.sheetnames[0]]

        return ws

    except EntitiesForbidden as err:
        report_error_to_rollbar(request=request, file=file, exception=err)
        raise ValidationError('This file may be damaged and '
                              'cannot be processed safely')


def format_percent(value):
    if value is None or isinstance(value, str):
        return value
    if value <= 1.0:
        str_value = str(value * 100)
    else:
        str_value = str(value)
    if str_value[-2:] == '.0':
        str_value = str_value[0:len(str_value)-2]
    return str_value + '%'


def format_cell_value(value):
    if value is None:
        return ''
    return str(value)


def parse_array_values(values):
    return title_array_values(
        remove_empty_array_values(
            strip_array_values(values)))


def strip_array_values(values):
    return [s.strip() for s in values]


def remove_empty_array_values(values):
    return [s for s in values if s]


def title_array_values(values):
    return [s.title() for s in values]


def parse_xlsx(file, request):
    try:
        ws = get_xlsx_sheet(file, request)

        # openpyxl package is 1-indexed
        percent_col = [get_column_letter(cell.column)
                       for cell in ws[2] if '%' in cell.number_format]

        if percent_col:
            for col in percent_col:
                for cell in ws[col]:
                    if cell.row != 1:
                        cell.value = format_percent(cell.value)

        ws_rows = ws.rows
        # Useing `next` will consome the row so that iteration of the data rows
        # will skipt the header row
        first_row = next(ws_rows)
        header = ','.join(
            [format_cell_value(cell.value) for cell in first_row])

        def format_row(row):
            return '"{}"'.format(
                '","'.join([format_cell_value(cell.value) for cell in row]))

        rows = [format_row(row)
                for row in ws_rows
                if any(cell.value is not None for cell in row)]

        return header, rows
    except Exception as err:
        report_error_to_rollbar(request=request, file=file, exception=err)
        raise ValidationError('Error parsing Excel (.xlsx) file')


def parse_csv(file, request):
    rows = []

    try:
        header = file.readline().decode(encoding='utf-8-sig').rstrip()
    except UnicodeDecodeError as err:
        report_error_to_rollbar(request=request, file=file, exception=err)
        raise ValidationError('Unsupported file encoding. Please '
                              'submit a UTF-8 CSV.')

    for idx, line in enumerate(file):
        if idx > 0:
            try:
                rows.append(line.decode(encoding='utf-8-sig').rstrip())
            except UnicodeDecodeError as err:
                report_error_to_rollbar(request=request,
                                        file=file,
                                        exception=err)
                raise ValidationError('Unsupported file encoding. Please '
                                      'submit a UTF-8 CSV.')

    return header, rows


def clean_row(row):
    invalid_keywords = ['N/A', 'n/a']
    replaced_value = replace_invalid_data(row, invalid_keywords)
    return cleanup_data(replaced_value)


def parse_csv_line(line):
    return list(csv.reader([line]))[0]


class ItemRemovedException(Exception):
    pass


@transaction.atomic
def parse_production_location_list(location_list: FacilityList):
    parsing_started = str(timezone.now())
    logger.info('[List Upload] Started CC Parse process!')

    print('Show path value >>>>')
    print(type(location_list.file.path))
    print(location_list.file.url)

    print('Show url value >>>>')
    print(type(location_list.file.path))
    print(location_list.file.path)

    print('Show DJANGO_ENV value >>>>')
    value = os.getenv('DJANGO_ENV', 'default_value')
    print(value)
    contri_cleaner = ContriCleaner(location_list.file, SectorCache())
    try:
        contri_cleaner_processed_data = contri_cleaner.process_data()
    except ParsingError as err:
        logger.error(f'[List Upload] Data Parsing Error: {err}')
        report_error_to_rollbar(
            message=str(err),
            extra_data=f'List file: {str(location_list)}'
            )
        raise ValidationError(str(err))  # TODO: save it to the DB
    except HandlerNotSetError as err:
        logger.error(f'[List Upload] Internal ContriCleaner Error: {err}')
        report_error_to_rollbar(
            message=str(err),
            extra_data=f'List file: {str(location_list)}'
            )
        raise Exception('Internal System Error. '
                        'Please contact support.')  # TODO: save it to the DB

    return
    processing_input = {
        'facility_list': location_list,
        'contri_cleaner_processed_data': contri_cleaner_processed_data,
        'parsing_started': parsing_started,
    }

    processing_facility_executor = ProcessingFacilityExecutor(
        ProcessingFacilityList(processing_input)
    )
    processing_facility_executor.run_processing()



# TODO: delete this function
def parse_facility_list_item_old(item):
    started = str(timezone.now())
    if type(item) != FacilityListItem:
        raise ValueError('Argument must be a FacilityListItem')
    if item.status == FacilityListItem.ITEM_REMOVED:
        raise ItemRemovedException
    if item.status != FacilityListItem.UPLOADED:
        raise ValueError('Items to be parsed must be in the UPLOADED status')
    try:
        is_geocoded = False
        fields = [f.lower()
                  for f in parse_csv_line(item.source.facility_list.header)]
        values = parse_csv_line(item.raw_data)

        # facility_type_processing_type is a special "meta" field that attempts
        # to simplify the submission process for contributors.
        if 'facility_type_processing_type' in fields:
            if 'facility_type' not in fields:
                fields.append('facility_type')
                values.append(
                    values[fields.index('facility_type_processing_type')])
            if 'processing_type' not in fields:
                fields.append('processing_type')
                values.append(
                    values[fields.index('facility_type_processing_type')])

        parser = CsvRowSectorProductTypeParser(fields, values)
        item.sector = parser.sectors
        if len(parser.product_types) > 0:
            if 'product_type' in fields:
                values[fields.index('product_type')] = parser.product_types
            else:
                fields.append('product_type')
                values.append(parser.product_types)
        else:
            # In this case the parse found that all of the product_type values
            # for the item were actually sectors. Setting None instead of an
            # empty list ensures that we do not create an extended field for
            # product_type
            if 'product_type' in fields:
                values[fields.index('product_type')] = None

        if CsvHeaderField.COUNTRY in fields:
            item.country_code = get_country_code(
                values[fields.index(CsvHeaderField.COUNTRY)])
        if CsvHeaderField.NAME in fields:
            item.name = values[fields.index(CsvHeaderField.NAME)]
            item.clean_name = clean(item.name)
            if item.clean_name is None:
                item.clean_name = ''
        if CsvHeaderField.ADDRESS in fields:
            item.address = values[fields.index(CsvHeaderField.ADDRESS)]
            item.clean_address = clean(item.address)
            if item.clean_address is None:
                item.clean_address = ''
        if CsvHeaderField.LAT in fields and CsvHeaderField.LNG in fields:
            lat = float(values[fields.index(CsvHeaderField.LAT)])
            lng = float(values[fields.index(CsvHeaderField.LNG)])
            item.geocoded_point = Point(lng, lat)
            is_geocoded = True

        create_extendedfields_for_listitem(item, fields, values)

        try:
            item.full_clean(exclude=('processing_started_at',
                                     'processing_completed_at',
                                     'processing_results', 'geocoded_point',
                                     'facility'))
            item.status = FacilityListItem.PARSED
            item.processing_results.append({
                'action': ProcessingAction.PARSE,
                'started_at': started,
                'error': False,
                'finished_at': str(timezone.now()),
                'is_geocoded': is_geocoded,
            })
        except ValidationError as ve:
            messages = []
            for name, errors in ve.error_dict.items():
                # We need to clear the invalid value so we can save the row
                setattr(item, name, '')
                error_str = ''.join(''.join(e.messages) for e in errors)
                messages.append(
                    'There is a problem with the {0}: {1}'.format(name,
                                                                  error_str)
                )

            # If there is a validation error on an array field, `full_clean`
            # appears to set it to an empty string which then causes `save`
            # to raise an exception.
            for field in {'sector'}:
                field_is_valid = (
                    getattr(item, field) is None
                    or isinstance(getattr(item, field), list))
                if not field_is_valid:
                    setattr(item, field, [])

            item.status = FacilityListItem.ERROR_PARSING
            item.processing_results.append({
                'action': ProcessingAction.PARSE,
                'started_at': started,
                'error': True,
                'message': '\n'.join(messages),
                'trace': traceback.format_exc(),
                'finished_at': str(timezone.now()),
            })
            logger.error(f'[List Upload] Parsing Error: {str(ve)}')
            logger.info(f'[List Upload] FacilityListItem Id: {item.id}')
    except Exception as e:
        item.status = FacilityListItem.ERROR_PARSING
        item.processing_results.append({
            'action': ProcessingAction.PARSE,
            'started_at': started,
            'error': True,
            'message': str(e),
            'trace': traceback.format_exc(),
            'finished_at': str(timezone.now()),
        })
        logger.error(f'[List Upload] Parsing Error: {str(e)}')
        logger.info(f'[List Upload] FacilityListItem Id: {item.id}')


def geocode_facility_list_item(item):
    started = str(timezone.now())
    if type(item) != FacilityListItem:
        raise ValueError('Argument must be a FacilityListItem')
    if item.status == FacilityListItem.ITEM_REMOVED:
        raise ItemRemovedException()
    if item.status != FacilityListItem.PARSED:
        raise ValueError('Items to be geocoded must be in the PARSED status')
    try:
        if item.geocoded_point is None:
            data = geocode_address(item.address, item.country_code)
            if data['result_count'] > 0:
                item.status = FacilityListItem.GEOCODED
                item.geocoded_point = Point(
                    data["geocoded_point"]["lng"],
                    data["geocoded_point"]["lat"]
                )
                item.geocoded_address = data["geocoded_address"]
            else:
                item.status = FacilityListItem.GEOCODED_NO_RESULTS
            item.processing_results.append({
                'action': ProcessingAction.GEOCODE,
                'started_at': started,
                'error': False,
                'skipped_geocoder': False,
                'data': data['full_response'],
                'finished_at': str(timezone.now()),
               })
        else:
            item.status = FacilityListItem.GEOCODED
            item.geocoded_address = item.address
            item.processing_results.append({
                'action': ProcessingAction.GEOCODE,
                'started_at': started,
                'error': False,
                'skipped_geocoder': True,
                'finished_at': str(timezone.now()),
            })

    except Exception as e:
        item.status = FacilityListItem.ERROR_GEOCODING
        item.processing_results.append({
            'action': ProcessingAction.GEOCODE,
            'started_at': started,
            'error': True,
            'message': str(e),
            'trace': traceback.format_exc(),
            'finished_at': str(timezone.now()),
        })
        logger.error(f'[List Upload] Geocoding Error: {str(e)}')
        logger.info(f'[List Upload] FacilityListItem Id: {item.id}')
        logger.info(f'[List Upload] Address: {item.address}, '
                    f'Country_Code: {item.country_code}')


def reduce_matches(matches):
    """
    Process a list of facility match scores to remove duplicate facilities,
    choosing the highest match score in the case of a duplicate.

    Arguments:
    matches -- A list of tuples of the format (extended_facility_id, score).

    Returns:
    A list of tuples in the format (facility_id, score). Extended facility

    Example:
        Input:
            [
                (US2020052GKF19F, 75),
                (US2020052GKF19F_MATCH-23, 88),
                (US2020052YDVKBQ, 45)
            ]
        Output:
            [
                (US2020052GKF19F, 88),
                (US2020052YDVKBQ, 45)
            ]
    """
    match_dict = {}
    for extended_id, score in matches:
        facility_id = normalize_extended_facility_id(extended_id)
        if facility_id not in match_dict or match_dict[facility_id] < score:
            match_dict[facility_id] = score
    return list(match_dict.items())


def is_string_match(item, facility):
    """
    Check if a list item is an exact string match to a facility, after
    processing both through the same string cleaning operations used by the
    matcher.

    Arguments:
    item -- A `FacilityListItem` instance being considered as a potential match
            to the specified facility.
    facility -- A `Facility` instance.

    Returns:
    True if the item is a string match to the facility
    """
    return (item.country_code == facility.country_code
            and clean(item.name) == clean(facility.name)
            and clean(item.address) == clean(facility.address))


def save_match_details(match_results, text_only_matches=None):
    """
    Save the results of a call to match_facility_list_items by creating
    Facility and FacilityMatch instances and updating the state of the affected
    FacilityListItems.

    Should be called in a transaction to ensure that all the updates are
    applied atomically.

    Arguments:
    match_results -- The dict return value from a call to
                     match_facility_list_items.
    text_only_matches -- An optional object where keys are FacilityListItem
                         IDs and values are a list of facilities that were
                         matched to the item without dedupe which should be
                         saved as pending matches.

    Returns:
    The list of `FacilityMatch` objects created
    """
    if text_only_matches is None:
        text_only_matches = {}
    processed_list_item_ids = match_results['processed_list_item_ids']
    item_matches = match_results['item_matches']
    results = match_results['results']
    started = match_results['started']
    finished = match_results['finished']

    automatic_threshold = results['automatic_threshold']

    def make_pending_match(item_id, facility_id, score):
        return FacilityMatch(
            facility_list_item_id=item_id,
            facility_id=facility_id,
            confidence=score,
            status=FacilityMatch.PENDING,
            results=results)

    all_matches = []
    for item_id, matches in item_matches.items():
        item = FacilityListItem.objects.get(id=item_id)
        item.status = FacilityListItem.POTENTIAL_MATCH
        matches = [make_pending_match(item_id, facility_id, score.item())
                   for facility_id, score in reduce_matches(matches)]

        if len(matches) == 1:
            if matches[0].confidence >= automatic_threshold:
                matches[0].status = FacilityMatch.AUTOMATIC
                matches[0].results['match_type'] = 'single_gazetteer_match'
                item.status = FacilityListItem.MATCHED
                item.facility = matches[0].facility
        else:
            quality_matches = [m for m in matches
                               if m.confidence > automatic_threshold]
            if len(quality_matches) == 1:
                matches[0].status = FacilityMatch.AUTOMATIC
                matches[0].results['match_type'] = \
                    'one_gazetteer_match_greater_than_threshold'
                item.status = FacilityListItem.MATCHED
                item.facility = matches[0].facility
            elif len(quality_matches) > 1:
                exact_matches = [m for m in quality_matches
                                 if is_string_match(item, m.facility)]
                # We check == 1 because multiple exact matches should not
                # happen. They are an indication of duplicate facility data
                # that should be merged through moderation tools. Showing the
                # multiple potential matches to the contributor increases the
                # visibility of the issue.
                if len(exact_matches) == 1:
                    exact_matches[0].status = FacilityMatch.AUTOMATIC
                    exact_matches[0].results['match_type'] = \
                        'multiple_gazetteer_matches_with_one_exact_string_match' # NOQA
                    item.status = FacilityListItem.MATCHED
                    item.facility = exact_matches[0].facility

        item.processing_results.append({
            'action': ProcessingAction.MATCH,
            'started_at': started,
            'error': False,
            'finished_at': finished
        })
        item.save()

        if item.source.create:
            for m in matches:
                m.save()

        all_matches.extend(matches)

    unmatched = (FacilityListItem.objects
                 .filter(id__in=processed_list_item_ids)
                 .exclude(id__in=item_matches.keys()))
    for item in unmatched:
        has_text_only_matches = (
            item.id in text_only_matches
            and len(text_only_matches[item.id]) > 0)
        if has_text_only_matches:
            text_only_results = copy.deepcopy(results)
            text_only_results['text_only_match'] = True
            text_only_match_objects = [
                FacilityMatch(
                    facility_list_item_id=item.id,
                    facility_id=facility.id,
                    confidence=0,
                    status=FacilityMatch.PENDING,
                    results=text_only_results)
                for facility in text_only_matches[item.id]]
            if item.source.create:
                for m in text_only_match_objects:
                    m.save()
            all_matches.extend(text_only_match_objects)

            item.status = FacilityListItem.POTENTIAL_MATCH
            item.processing_results.append({
                'action': ProcessingAction.MATCH,
                'started_at': started,
                'error': False,
                'text_only_match': True,
                'finished_at': finished
            })
        elif item.status == FacilityListItem.GEOCODED_NO_RESULTS:
            item.status = FacilityListItem.ERROR_MATCHING
            item.processing_results.append({
                'action': ProcessingAction.MATCH,
                'started_at': started,
                'error': True,
                'message': ('No match to an existing facility and cannot '
                            'create a new facility without a geocode result'),
                'finished_at': finished
            })
        else:
            if item.source.create:
                facility = Facility(name=item.name,
                                    address=item.address,
                                    country_code=item.country_code,
                                    location=item.geocoded_point,
                                    created_from=item)
                facility.save()

                match = make_pending_match(item.id, facility.id, 1.0)
                match.results['match_type'] = 'no_gazetteer_match'
                match.status = FacilityMatch.AUTOMATIC
                match.save()

                item.facility = facility
            item.status = FacilityListItem.MATCHED
            item.processing_results.append({
                'action': ProcessingAction.MATCH,
                'started_at': started,
                'error': False,
                'finished_at': finished
            })
        item.save()

    items = FacilityListItem.objects.filter(id__in=processed_list_item_ids) \
                                    .exclude(facility__isnull=True)
    for item in items:
        update_extendedfields_for_list_item(item)

    return all_matches


def save_exact_match_details(exact_results):
    """
    Save the results of a call to identify_exact_matches by creating
    Facility and FacilityMatch instances and updating the state of the affected
    FacilityListItems.

    Should be called in a transaction to ensure that all the updates are
    applied atomically.

    Arguments:
    exact_results -- The dict return value from a call to
                    identify_exact_matches

    Returns:
    The list of `FacilityMatch` objects created
    """
    processed_list_item_ids = exact_results['processed_list_item_ids']
    item_matches = exact_results['item_matches']
    started = exact_results['started']
    finished = exact_results['finished']

    def make_pending_match(item_id, facility_id):
        return FacilityMatch(
            facility_list_item_id=item_id,
            facility_id=facility_id,
            confidence=1.0,
            status=FacilityMatch.PENDING,
            results={})

    all_matches = []
    for item_id, exact_matches in item_matches.items():
        item = FacilityListItem.objects.get(id=item_id)
        item.status = FacilityListItem.POTENTIAL_MATCH

        matches = [make_pending_match(item_id, m.get('facility_id'))
                   for m in exact_matches]
        matches[0].status = FacilityMatch.AUTOMATIC
        item.status = FacilityListItem.MATCHED
        item.facility_id = matches[0].facility_id

        if len(matches) == 1:
            matches[0].results['match_type'] = 'single_exact_match'
        else:
            matches[0].results['match_type'] = 'multiple_exact_matches'

        item.processing_results.append({
            'action': ProcessingAction.MATCH,
            'started_at': started,
            'error': False,
            'finished_at': finished,
            'exact_match': True
        })
        item.save()

        if item.source.create:
            for m in matches:
                m.save()

        all_matches.extend(matches)

    items = FacilityListItem.objects.filter(id__in=processed_list_item_ids) \
                                    .exclude(facility__isnull=True)
    for item in items:
        update_extendedfields_for_list_item(item)

    return all_matches


def handle_external_match_process_result(id, result, request, should_create):
    context = {'request': request}
    if should_create:
        list_item_object_type = FacilityListItem
    else:
        list_item_object_type = FacilityListItemTemp
    end_match_process_status = [list_item_object_type.MATCHED,
                                list_item_object_type.POTENTIAL_MATCH,
                                list_item_object_type.ERROR_MATCHING]
    f_l_item = None
    timer = 0
    timeout = 25
    while True:
        if timer > timeout:
            break
        facility_list_item = list_item_object_type.objects.get(id=id)
        if facility_list_item.status in end_match_process_status:
            f_l_item = facility_list_item
            break
        time.sleep(1)
        timer = timer + 1

    if f_l_item is None:
        # Timeout
        return get_error_match_result(id, result)
    if should_create:
        match_object_type = FacilityMatch
    else:
        match_object_type = FacilityMatchTemp
    queryset_f_m = match_object_type.objects.filter(
        facility_list_item=f_l_item.id)
    if queryset_f_m.count() == 0:
        # No Match and Geocoder Returned No Results
        return get_error_match_result(f_l_item.id, result)

    automatic_matches = [
        match for match in queryset_f_m
        if match.status == match_object_type.AUTOMATIC
    ]
    if (len(automatic_matches) == 1):
        if isinstance(f_l_item, FacilityListItem):
            # New Facility
            if f_l_item.facility is None:
                return get_new_facility_match_result(f_l_item.id, None, result)
            if f_l_item.facility.created_from == f_l_item:
                return get_new_facility_match_result(
                    f_l_item.id, f_l_item.facility.id, result
                )

            # Automatic Match
            return get_automatic_match_result(f_l_item.id,
                                              f_l_item.facility.id,
                                              automatic_matches[0].confidence,
                                              context,
                                              result)
        else:
            # New Facility
            if f_l_item.facility_id is None:
                return get_new_facility_match_result(f_l_item.id, None, result)

            # Automatic Match
            return get_automatic_match_result(f_l_item.id,
                                              f_l_item.facility_id,
                                              automatic_matches[0].confidence,
                                              context,
                                              result)

    pending_matches = [
        match for match in queryset_f_m
        if match.status == match_object_type.PENDING
    ]
    for item in pending_matches:
        # Potential Match
        return get_potential_match_result(f_l_item,
                                          item,
                                          len(pending_matches),
                                          context,
                                          should_create,
                                          result)

    return result


def get_error_match_result(id, result):
    result['status'] = FacilityListItem.ERROR_MATCHING
    result['item_id'] = id

    return result


def get_new_facility_match_result(list_item_id, facility_id, result):
    result['status'] = FacilityListItem.NEW_FACILITY
    result['item_id'] = list_item_id

    if facility_id is not None:
        result['os_id'] = facility_id

    return result


def get_automatic_match_result(list_item_id, facility_id, confidence, context,
                               result):
    from api.serializers import FacilityIndexDetailsSerializer

    facility_index = FacilityIndex.objects.get(id=facility_id)
    facility_dict = FacilityIndexDetailsSerializer(
        facility_index,
        context=context
        ).data
    facility_dict['confidence'] = confidence

    result['matches'].append(facility_dict)
    result['item_id'] = list_item_id
    result['os_id'] = facility_id
    result['status'] = FacilityListItem.MATCHED

    return result


def get_potential_match_result(list_item, item, items_count, context,
                               should_create, result):
    from api.serializers import FacilityIndexDetailsSerializer

    AUTOMATIC_THRESHOLD = 0.8

    facility = Facility.objects.get(id=item.facility_id)
    facility_index = FacilityIndex.objects.get(id=facility.id)
    facility_dict = FacilityIndexDetailsSerializer(
        facility_index,
        context=context
        ).data
    facility_dict['confidence'] = item.confidence

    if item.confidence < AUTOMATIC_THRESHOLD or items_count > 1:
        if should_create:
            facility_dict['confirm_match_url'] = reverse(
                'facility-match-confirm',
                kwargs={'pk': item.id})
            facility_dict['reject_match_url'] = reverse(
                'facility-match-reject',
                kwargs={'pk': item.id})
    result['matches'].append(facility_dict)
    result['item_id'] = list_item.id
    result['status'] = list_item.status

    return result
