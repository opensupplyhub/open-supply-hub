"""Unit tests for :mod:`contribot`.

The suite constructs real :class:`contribot.ContriBot` instances from
small temporary workbooks (see ``conftest.py``) and exercises each checker /
helper in isolation. No database or network access happens; ``save`` is tested
against an in-memory SQLite engine.
"""

import openpyxl
import pandas as pd
import pytest
from sqlalchemy import create_engine

import contribot
from utils import map_n_dataframe_cols_to_excel_cols


# --------------------------------------------------------------------------- #
# utils / column mapping
# --------------------------------------------------------------------------- #
class TestColumnLetterMapping:
    def test_single_letters(self):
        assert map_n_dataframe_cols_to_excel_cols(3) == ["A", "B", "C"]

    def test_full_alphabet_boundary(self):
        cols = map_n_dataframe_cols_to_excel_cols(28)
        assert cols[25] == "Z"
        assert cols[26] == "AA"
        assert cols[27] == "AB"

    def test_length_matches_request(self):
        assert len(map_n_dataframe_cols_to_excel_cols(100)) == 100


# --------------------------------------------------------------------------- #
# construction / workbook layout
# --------------------------------------------------------------------------- #
class TestConstruction:
    def test_helper_sheets_created_in_order(self, contribution):
        names = contribution.wb.sheetnames
        assert names[:3] == ["Summary", "Findings", "Similarities"]
        assert contribution.fixessheet in names

    def test_target_filename_marks_processed(self, contribution):
        assert contribution.targetfilename == "contribution.~PROCESSED.xlsx"

    def test_column_names_are_lowercased(self, contribution):
        assert list(contribution.df.columns) == [
            "country",
            "name",
            "address",
            "sector_product_type",
        ]

    def test_dataframe_index_offset_by_two(self, contribution):
        assert list(contribution.df.index) == [2, 3]

    def test_column_to_excel_mapping(self, contribution):
        assert contribution.mapping["country"] == "A"
        assert contribution.mapping["sector_product_type"] == "D"

    def test_findings_sheet_has_header_and_values(self, contribution):
        ws = contribution.wb["Findings"]
        assert ws["A1"].value == "country"
        assert ws["A2"].value == "Bangladesh"

    def test_extra_source_sheets_flagged(self, tmp_path, config_path, monkeypatch):
        monkeypatch.chdir(tmp_path)
        path = tmp_path / "multi.xlsx"
        with pd.ExcelWriter(path) as writer:
            pd.DataFrame({"country": ["Germany"]}).to_excel(
                writer, sheet_name="Data", index=False
            )
            pd.DataFrame({"x": [1]}).to_excel(writer, sheet_name="Notes", index=False)
        bot = contribot.ContriBot(str(path), config_file=config_path)
        assert any(d["code"] == "T0015" for d in bot.diagnostics_table)


# --------------------------------------------------------------------------- #
# configuration loading
# --------------------------------------------------------------------------- #
class TestGetConfig:
    def test_config_loaded_from_explicit_path(self, contribution):
        assert contribution.have_config is True
        assert contribution.df_config_string.startswith("GDRIVE")
        assert "C0001" in set(contribution.df_config.code)

    def test_missing_config_yields_empty_definitions(self, make_contribution):
        osh = make_contribution(with_config=False)
        assert osh.have_config is False
        assert len(osh.df_config) == 0
        assert osh.df_config_string == "No valid configuration found"


# --------------------------------------------------------------------------- #
# stats / metrics accumulators
# --------------------------------------------------------------------------- #
class TestStatsAndMetrics:
    def test_add_stats_appends_record(self, contribution):
        contribution.statistics.clear()
        contribution._add_stats(level="table", name="t", metric="m", value=3)
        assert contribution.statistics == [
            {"level": "table", "name": "t", "metric": "m", "value": 3}
        ]

    def test_add_metrics_appends_record(self, contribution):
        contribution.metrics.clear()
        contribution._add_metrics(level="column", name="country", metric="m", value=1)
        assert contribution.metrics[-1]["name"] == "country"


# --------------------------------------------------------------------------- #
# config lookup helpers
# --------------------------------------------------------------------------- #
class TestConfigLookups:
    def test_get_severity_known(self, contribution):
        assert contribution._get_severity("C0001") == 100

    def test_get_severity_unknown_returns_minus_one(self, contribution):
        assert contribution._get_severity("NOPE9") == -1

    def test_decide_autofix_true(self, contribution):
        assert bool(contribution._decide_autofix("C0008")) is True

    def test_decide_autofix_false(self, contribution):
        assert bool(contribution._decide_autofix("C0016")) is False

    def test_decide_autofix_missing_code(self, contribution):
        assert contribution._decide_autofix("ZZZZ") is False

    def test_style_findings_valid(self, contribution):
        assert contribution._get_style_findings("C0001") == "Bad"

    def test_style_findings_blank_when_not_allowed(self, contribution):
        assert contribution._get_style_findings("C0007") == ""

    def test_style_fixed_valid(self, contribution):
        assert contribution._get_style_fixed("C0008") == "Good"

    def test_style_fixed_blank_default(self, contribution):
        assert contribution._get_style_fixed("C0013") == ""

    def test_highest_severity_style_fixed(self, contribution):
        # C0011 (severity 30, "Bad") outranks C0010 (severity 5, "Good").
        assert contribution._get_highest_severity_style_fixed(["C0010", "C0011"]) == "Bad"

    def test_tokens_above_threshold(self, contribution):
        assert contribution._tokens_above_threshold(("x", 2)) is True
        assert contribution._tokens_above_threshold(("x", 1)) is False


# --------------------------------------------------------------------------- #
# _add_diagnosis
# --------------------------------------------------------------------------- #
class TestAddDiagnosis:
    def test_table_code_goes_to_table_diagnostics(self, contribution):
        contribution._add_diagnosis(code="T0001")
        rec = contribution.diagnostics_table[-1]
        assert rec["level"] == "table"
        assert rec["severity"] == 100
        assert rec["error"] == "missing country column"

    def test_column_code_builds_cell_reference(self, contribution):
        contribution._add_diagnosis(code="C0001", column_name="country", row=2,
                                    country="neverland")
        rec = contribution.diagnostics_column[-1]
        assert rec["level"] == "column"
        assert rec["cell"] == "A2"
        assert rec["error"] == "country 'neverland' is not recognised"

    def test_unknown_column_name_uses_at_fallback(self, contribution):
        contribution._add_diagnosis(code="C0001", column_name="not_a_column", row=7,
                                    country="x")
        assert contribution.diagnostics_column[-1]["cell"] == "@7"

    def test_unknown_code_reports_tool_error(self, contribution):
        contribution._add_diagnosis(code="C9999", column_name="country", row=2)
        rec = contribution.diagnostics_column[-1]
        assert rec["severity"] == 100
        assert "Tool error" in rec["error"]

    def test_template_renders_all_supplied_params(self, contribution):
        contribution._add_diagnosis(code="C0006", column_name="address", row=3,
                                    country_in_address="germany", country="bangladesh")
        error = contribution.diagnostics_column[-1]["error"]
        assert error == ("country germany in address, "
                         "country column says bangladesh")

    def test_template_flags_missing_config_variable(self, contribution):
        contribution._add_diagnosis(code="C0006", column_name="address", row=3,
                                    country="bangladesh")
        error = contribution.diagnostics_column[-1]["error"]
        assert "#CfgErr" in error
        assert "bangladesh" in error


# --------------------------------------------------------------------------- #
# _add_fix_comment
# --------------------------------------------------------------------------- #
class TestAddFixComment:
    def test_appends_when_autofix_enabled(self, contribution):
        contribution._add_fix_comment(coordinates="A2", code="C0008")
        assert contribution.fixes[-1] == {
            "coordinates": "A2",
            "fix": "removed double quotes",
        }

    def test_noop_when_autofix_disabled(self, contribution):
        before = len(contribution.fixes)
        contribution._add_fix_comment(coordinates="A2", code="C0016")
        assert len(contribution.fixes) == before


# --------------------------------------------------------------------------- #
# normalise_column_names
# --------------------------------------------------------------------------- #
def test_normalise_column_names(contribution):
    df = pd.DataFrame({"Foo": [1], "BAR": [2]})
    out = contribution.normalise_column_names(df)
    assert list(out.columns) == ["foo", "bar"]


# --------------------------------------------------------------------------- #
# check_table
# --------------------------------------------------------------------------- #
class TestCheckTable:
    def _codes(self, osh):
        return {d["code"] for d in osh.diagnostics_table}

    def test_records_row_and_column_counts(self, contribution):
        contribution.check_table()
        metrics = {(s["metric"], s["value"]) for s in contribution.statistics}
        assert ("num_rows", 2) in metrics
        assert ("num_columns", 4) in metrics

    def test_missing_required_columns(self, make_contribution):
        df = pd.DataFrame({"name": ["Acme Textiles Ltd"], "address": ["123 Road City"]})
        osh = make_contribution(df=df)
        osh.check_table()
        codes = self._codes(osh)
        assert {"T0001", "T0004"} <= codes  # missing country + sector
        assert "T0002" not in codes  # name present
        assert "T0003" not in codes  # address present

    def test_lat_without_lng(self, make_contribution):
        df = pd.DataFrame({"country": ["Germany"], "lat": [52.5]})
        osh = make_contribution(df=df)
        osh.check_table()
        assert "T0005" in self._codes(osh)

    def test_lng_without_lat(self, make_contribution):
        df = pd.DataFrame({"country": ["Germany"], "lng": [13.4]})
        osh = make_contribution(df=df)
        osh.check_table()
        assert "T0006" in self._codes(osh)

    def test_both_lat_and_lng(self, make_contribution):
        df = pd.DataFrame({"country": ["Germany"], "lat": [52.5], "lng": [13.4]})
        osh = make_contribution(df=df)
        osh.check_table()
        assert "T0007" in self._codes(osh)

    def test_unrecognised_column_flagged(self, make_contribution):
        df = pd.DataFrame({"country": ["Germany"], "surprise": ["x"]})
        osh = make_contribution(df=df)
        osh.check_table()
        assert "T0008" in self._codes(osh)

    def test_empty_table_flagged(self, make_contribution):
        df = pd.DataFrame(
            {"country": [], "name": [], "address": [], "sector_product_type": []}
        )
        osh = make_contribution(df=df)
        osh.check_table()
        assert "T0009" in self._codes(osh)


# --------------------------------------------------------------------------- #
# check_column_country
# --------------------------------------------------------------------------- #
class TestCheckColumnCountry:
    def test_known_countries_pass(self, contribution):
        contribution.check_column_country()
        assert not any(d["code"] == "C0001" for d in contribution.diagnostics_column)

    def test_unknown_country_flagged(self, make_contribution):
        df = pd.DataFrame(
            {
                "country": ["Neverland"],
                "name": ["Acme Textiles Ltd"],
                "address": ["123 Road City"],
            }
        )
        osh = make_contribution(df=df)
        osh.check_column_country()
        recs = [d for d in osh.diagnostics_column if d["code"] == "C0001"]
        assert len(recs) == 1
        assert "neverland" in recs[0]["error"]

    def test_missing_country_column(self, make_contribution):
        df = pd.DataFrame({"name": ["Acme Textiles Ltd"], "address": ["123 Road City"]})
        osh = make_contribution(df=df)
        osh.check_column_country()
        assert any(d["code"] == "C0002" for d in osh.diagnostics_column)


# --------------------------------------------------------------------------- #
# check_for_country_name_in_address
# --------------------------------------------------------------------------- #
class TestCountryNameInAddress:
    def test_mismatched_long_country_name(self, make_contribution):
        df = pd.DataFrame(
            {"country": ["Bangladesh"], "name": ["A"], "address": ["big road germany"]}
        )
        osh = make_contribution(df=df)
        osh.check_for_country_name_in_address()
        assert any(d["code"] == "C0006" for d in osh.diagnostics_column)

    def test_mismatched_short_country_code(self, make_contribution):
        df = pd.DataFrame(
            {"country": ["Bangladesh"], "name": ["A"], "address": ["big road are"]}
        )
        osh = make_contribution(df=df)
        osh.check_for_country_name_in_address()
        assert any(d["code"] == "C0025" for d in osh.diagnostics_column)

    def test_matching_country_no_warning(self, make_contribution):
        df = pd.DataFrame(
            {"country": ["China"], "name": ["A"], "address": ["some road china"]}
        )
        osh = make_contribution(df=df)
        osh.check_for_country_name_in_address()
        assert not any(
            d["code"] in {"C0006", "C0025"} for d in osh.diagnostics_column
        )


# --------------------------------------------------------------------------- #
# single-column checkers
# --------------------------------------------------------------------------- #
def _single_col(make_contribution, column, values):
    df = pd.DataFrame({column: values})
    return make_contribution(df=df)


class TestUnusualCharacters:
    def test_flags_non_latin(self, make_contribution):
        osh = _single_col(make_contribution, "name", ["Acme \u5de5\u5382"])
        osh.check_column_unusual_characters("name")
        assert any(d["code"] == "C0007" for d in osh.diagnostics_column)

    def test_plain_latin_is_clean(self, make_contribution):
        osh = _single_col(make_contribution, "name", ["Acme Textiles Ltd"])
        osh.check_column_unusual_characters("name")
        assert not any(d["code"] == "C0007" for d in osh.diagnostics_column)


class TestDoubleQuotes:
    def test_flags_and_autofixes(self, make_contribution):
        osh = _single_col(make_contribution, "name", ['He said "hi" there'])
        osh.check_column_double_quotes("name")
        assert any(d["code"] == "C0008" for d in osh.diagnostics_column)
        fixed = osh.wb[osh.fixessheet]["A2"].value
        assert '"' not in fixed

    def test_no_quotes_no_finding(self, make_contribution):
        osh = _single_col(make_contribution, "name", ["Clean value"])
        osh.check_column_double_quotes("name")
        assert not any(d["code"] == "C0008" for d in osh.diagnostics_column)


class TestPoBox:
    @pytest.mark.parametrize("value", ["P.O. Box 12", "PO Box 9", "Post Office 3"])
    def test_flags_po_box(self, make_contribution, value):
        osh = _single_col(make_contribution, "address", [value])
        osh.check_column_po_box("address")
        assert any(d["code"] == "C0009" for d in osh.diagnostics_column)

    def test_regular_address_clean(self, make_contribution):
        osh = _single_col(make_contribution, "address", ["221B Baker Street"])
        osh.check_column_po_box("address")
        assert not any(d["code"] == "C0009" for d in osh.diagnostics_column)


class TestLeadingTrailingBlanks:
    @pytest.mark.parametrize(
        "value,code",
        [
            (" leading", "C0010"),
            ("trailing ", "C0011"),
            ("double  space", "C0012"),
            ("tab\there", "C0013"),
            ("carriage\rreturn", "C0014"),
            ("new\nline", "C0015"),
        ],
    )
    def test_each_whitespace_variant(self, make_contribution, value, code):
        osh = _single_col(make_contribution, "name", ["placeholder"])
        # Assign in-memory: Excel normalises bare control chars (e.g. \r) on the
        # file round-trip, so inject the raw value the checker actually reads.
        osh.df["name"] = [value]
        osh.check_column_leading_trailing_blanks("name")
        assert any(d["code"] == code for d in osh.diagnostics_column)

    def test_autofix_trims_value(self, make_contribution):
        osh = _single_col(make_contribution, "name", ["  messy   value  "])
        osh.check_column_leading_trailing_blanks("name")
        assert osh.wb[osh.fixessheet]["A2"].value == "messy value"

    def test_clean_value_untouched(self, make_contribution):
        osh = _single_col(make_contribution, "name", ["tidy value"])
        osh.check_column_leading_trailing_blanks("name")
        assert not osh.diagnostics_column


class TestColumnLengths:
    def test_too_long(self, make_contribution):
        osh = _single_col(make_contribution, "name", ["x" * 250])
        osh.check_column_lengths("name")
        assert any(d["code"] == "C0016" for d in osh.diagnostics_column)

    def test_name_too_short(self, make_contribution):
        osh = _single_col(make_contribution, "name", ["Acme"])
        osh.check_column_lengths("name")
        assert any(d["code"] == "C0017" for d in osh.diagnostics_column)

    def test_address_too_short(self, make_contribution):
        osh = _single_col(make_contribution, "address", ["short addr"])
        osh.check_column_lengths("address")
        assert any(d["code"] == "C0018" for d in osh.diagnostics_column)

    def test_reasonable_length_clean(self, make_contribution):
        osh = _single_col(make_contribution, "name", ["Acme Textiles Limited"])
        osh.check_column_lengths("name")
        assert not osh.diagnostics_column


class TestExcelErrorCodes:
    def test_flags_and_strips_token(self, make_contribution):
        osh = _single_col(make_contribution, "name", ["status N/A end"])
        osh.check_column_excel_error_codes_na("name")
        assert any(d["code"] == "C0019" for d in osh.diagnostics_column)
        assert "N/A" not in osh.wb[osh.fixessheet]["A2"].value

    def test_clean_value(self, make_contribution):
        osh = _single_col(make_contribution, "name", ["all good here"])
        osh.check_column_excel_error_codes_na("name")
        assert not osh.diagnostics_column


class TestMultipleCommas:
    def test_flags_and_autofixes(self, make_contribution):
        osh = _single_col(make_contribution, "name", ["alpha,,beta"])
        osh.check_column_multiple_commas("name")
        assert any(d["code"] == "C0020" for d in osh.diagnostics_column)
        assert osh.wb[osh.fixessheet]["A2"].value == "alpha, beta"

    def test_single_comma_clean(self, make_contribution):
        osh = _single_col(make_contribution, "name", ["alpha, beta"])
        osh.check_column_multiple_commas("name")
        assert not osh.diagnostics_column


class TestCorrectlyDelimited:
    def test_comma_only_single_field(self, make_contribution):
        osh = _single_col(make_contribution, "sector_product_type", ["cotton,wool"])
        osh.check_column_correctly_delimited("sector_product_type")
        assert any(d["code"] == "C0022" for d in osh.diagnostics_column)

    def test_comma_inside_pipe_field(self, make_contribution):
        osh = _single_col(
            make_contribution, "sector_product_type", ["cotton,wool|silk,linen"]
        )
        osh.check_column_correctly_delimited("sector_product_type")
        assert any(d["code"] == "C0021" for d in osh.diagnostics_column)

    def test_properly_pipe_delimited_clean(self, make_contribution):
        osh = _single_col(
            make_contribution, "sector_product_type", ["cotton|wool|silk"]
        )
        osh.check_column_correctly_delimited("sector_product_type")
        assert not any(
            d["code"] in {"C0021", "C0022", "C0023"} for d in osh.diagnostics_column
        )


# --------------------------------------------------------------------------- #
# _get_height_for_row
# --------------------------------------------------------------------------- #
class TestRowHeight:
    def _sheet(self, value):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions["A"].width = 10
        ws["A1"] = value
        return ws

    def test_short_value_uses_base_height(self, contribution):
        ws = self._sheet("short")
        assert contribution._get_height_for_row(ws, 0, font_size=12) == 16

    def test_long_value_grows_height(self, contribution):
        short = contribution._get_height_for_row(self._sheet("short"), 0)
        long = contribution._get_height_for_row(self._sheet("x" * 200), 0)
        assert long > short


# --------------------------------------------------------------------------- #
# check_name_address_duplicates
# --------------------------------------------------------------------------- #
class TestDuplicates:
    def test_single_row_returns_early(self, make_contribution):
        df = pd.DataFrame(
            {"country": ["Germany"], "name": ["Acme Ltd"], "address": ["123 Road City"]}
        )
        osh = make_contribution(df=df)
        osh.check_name_address_duplicates()
        assert osh.diagnostics_table == []

    def test_identical_rows_flagged(self, make_contribution):
        df = pd.DataFrame(
            {
                "country": ["Bangladesh", "Bangladesh"],
                "name": ["Acme Textiles Limited", "Acme Textiles Limited"],
                "address": [
                    "123 Industrial Road, Dhaka",
                    "123 Industrial Road, Dhaka",
                ],
            }
        )
        osh = make_contribution(df=df)
        osh.check_name_address_duplicates()
        assert osh.num_identical >= 1
        assert any(d["code"] == "T0010" for d in osh.diagnostics_table)
        assert len(osh.df_similarity) >= 1


# --------------------------------------------------------------------------- #
# check_columns orchestration
# --------------------------------------------------------------------------- #
def test_check_columns_runs_end_to_end(contribution):
    contribution.check_columns()
    # Default frame is clean; the orchestrator should complete without raising
    # and populate the similarity frame.
    assert hasattr(contribution, "df_similarity")


# --------------------------------------------------------------------------- #
# populate_summary / save
# --------------------------------------------------------------------------- #
class TestSummaryAndSave:
    def test_populate_summary_builds_summary(self, contribution):
        contribution.check_table()
        summary = contribution.populate_summary(write_to_jsonl=False)
        assert set(summary) == {
            "sourcefilename",
            "targetfilename",
            "error_ratio",
            "num_lines",
            "num_errors",
        }
        assert summary["num_lines"] == 2

    def test_save_writes_file_and_returns_summary(self, contribution, tmp_path):
        contribution.check_table()
        engine = create_engine("sqlite://")
        target = tmp_path / "out"
        summary = contribution.save(
            targetfolder=str(target), engine=engine, write_to_jsonl=False
        )
        assert (target / contribution.targetfilename).exists()
        assert summary["sourcefilename"] == "contribution.xlsx"
