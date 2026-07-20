"""Validate Open Supply Hub facility list uploads and produce annotated workbooks.

ContriBot reads a contributor Excel workbook, runs table- and column-level quality
checks, applies optional auto-fixes, and writes an annotated output workbook with
Summary, Findings, Similarities, and Fixes sheets.

Error codes and severities come from an external configuration workbook
(``0000.error_codes.xlsx`` by default). Codes starting with ``T`` are table-level;
codes starting with ``C`` are column-level.
"""

import datetime
import os
import re
import string
import sys
import unicodedata
import warnings
from copy import copy
from math import ceil
from pathlib import Path

import cleanco
import openpyxl
import pandas as pd
import rapidfuzz
from jinja2.nativetypes import NativeEnvironment
from jinja2.nodes import TemplateData
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScale, FormatObject, Rule
from openpyxl.styles import Alignment, Color, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm
from unidecode import unidecode

from known_countries import COUNTRY_CODES, known_countries
from utils import map_n_dataframe_cols_to_excel_cols


class ContriBot:
    """Validate a facility list workbook and build an annotated output workbook.

    Typical usage::

        bot = ContriBot("contribution.xlsx")
        bot.process()
        bot.save(targetfolder="./output")

    Attributes:
        df: Normalised facility data (column names lowercased, index offset for Excel rows).
        wb: The openpyxl workbook, including generated report sheets.
        diagnostics_table: Table-level findings (``T*`` codes).
        diagnostics_column: Cell-level findings (``C*`` codes).
        summary: Run statistics populated by :meth:`populate_summary`.
    """

    def __init__(self, filename, config_file="0000.error_codes.xlsx"):
        """Load a workbook and prepare report sheets.

        Args:
            filename: Path to the contributor ``.xlsx`` file.
            config_file: Path to the error-codes configuration workbook.
        """
        warnings.filterwarnings("ignore", module="openpyxl")
        self.START = datetime.datetime.now()
        self.sourcepath, self.sourcefilename = os.path.split(filename)
        parts = self.sourcefilename.split(".")
        self.targetfilename = parts[0] + ".~PROCESSED." + ".".join(parts[1:])

        self.wb = openpyxl.load_workbook(filename)

        self.sourcesheet = self.wb.sheetnames[0]
        if len(self.wb._sheets) > 1:
            existing_tables = [s for s in self.wb.sheetnames]
            residual = list(range(len(self.wb._sheets)))
        else:
            existing_tables = [s for s in self.wb.sheetnames]
            residual = []
        self.wb.create_sheet("Summary")
        summary_sheet_no = len(self.wb._sheets) - 1
        self.wb.create_sheet("Findings")
        findings_sheet_no = len(self.wb._sheets) - 1
        self.wb.create_sheet("Similarities")
        similarities_sheet_no = len(self.wb._sheets) - 1
        self.fixessheet = f"{self.sourcesheet:.25s}~Fixes"
        self.wb.create_sheet(self.fixessheet)
        fixes_sheet_no = len(self.wb._sheets) - 1

        self.wb._sheets = [
            self.wb._sheets[summary_sheet_no],
            self.wb._sheets[findings_sheet_no],
            self.wb._sheets[similarities_sheet_no],
            self.wb._sheets[0],
            self.wb._sheets[fixes_sheet_no],
        ] + [self.wb._sheets[r] for r in residual[1:]]

        self.df = pd.read_excel(filename)
        if len(self.df) > 10000:
            self._add_diagnosis(code="T0016", num_lines=len(self.df))
            self.df = self.df[:10000]

        self.df.index = self.df.index + 2
        self.df = self.normalise_column_names(self.df)

        # Create a column mapping for excel, the earlier version could not handle files with more than 26 cols (len(en-alphabet))
        self.mapping = dict(
            zip(
                self.df.columns,
                map_n_dataframe_cols_to_excel_cols(len(self.df.columns)),
            )
        )

        self.copy_values_and_layout(self.wb, self.df, self.sourcesheet, "Findings")
        self.copy_values_and_layout(self.wb, self.df, self.sourcesheet, self.fixessheet)

        self.statistics = []
        self.metrics = []
        self.diagnostics_table = []
        self.diagnostics_column = []
        self.fixes = []
        self.comments = {}

        self.known_countries = known_countries

        self.config_file = config_file
        self._get_config()

        if len(existing_tables) > 1:
            self._add_diagnosis(code="T0015", existing_sheets=",".join(existing_tables))

    def _get_config(self):
        """Load error-code definitions from the config or fallback file.

        Sets ``self.df_config``, ``self.df_config_string``, and ``self.have_config``.
        Falls back to ``error_codes.xlsx`` in the working directory, then to an empty
        config if neither file exists.
        """
        if os.path.exists(self.config_file):
            self.df_config = pd.read_excel(
                self.config_file, sheet_name="Definitions"
            ).fillna("")
            config_modification_dt = datetime.datetime.fromtimestamp(
                os.path.getmtime(self.config_file)
            )
            self.df_config_string = (
                f"GDRIVE {config_modification_dt:%Y-%m-%d %H:%M:%S} UTC"
            )
            self.have_config = True
        elif os.path.exists("error_codes.xlsx"):
            self.df_config = pd.read_excel(
                "error_codes.xlsx", sheet_name="Definitions"
            ).fillna("")
            config_modification_dt = datetime.datetime.fromtimestamp(
                os.path.getmtime("error_codes.xlsx")
            )
            self.df_config_string = (
                f"DEFAULT {config_modification_dt:%Y-%m-%d %H:%M:%S} UTC"
            )
            self.have_config = True
        else:
            self.df_config = pd.DataFrame(
                {
                    "code": [],
                    "level": [],
                    "severity": [],
                    "style_findings": [],
                    "style_fixed": [],
                    "auto_fix": [],
                    "topic": [],
                    "error": [],
                    "corrective_action": [],
                    "fixes_comment": [],
                }
            )
            self.df_config_string = "No valid configuration found"
            self.have_config = False

    def process(self):
        """Run the full validation pipeline on the loaded workbook.

        Calls :meth:`check_table`, :meth:`check_columns`, and
        :meth:`_add_comments_to_excel_sheets` in order. Does not write output;
        call :meth:`save` afterward to produce the annotated workbook.
        """
        self.check_table()
        self.check_columns()
        self._add_comments_to_excel_sheets()

    def save(self, targetfolder="./sun"):
        """Write the processed workbook.

        Args:
            targetfolder: Directory for the output ``.~PROCESSED.xlsx`` file.

        Returns:
            dict: Summary with ``error_ratio``, ``num_lines``, ``num_errors``,
            and runtime fields.
        """
        Path(targetfolder).mkdir(exist_ok=True, parents=True)
        self.populate_summary()
        self.fixes_sheet_fontsize()
        filepath = f"{targetfolder}/{self.targetfilename}".replace("//", "/")
        print(
            f"save targetfolder {targetfolder} self.targetfilename {self.targetfilename}"
        )

        self.wb.active = self.wb["Summary"]
        self.wb.save(filepath)

        end = datetime.datetime.now()
        self.summary["datetime_started"] = self.START.isoformat()
        self.summary["datetime_completed"] = end.isoformat()
        self.summary["runtime_total_seconds"] = (end - self.START).total_seconds()

        return self.summary

    def fixes_sheet_fontsize(self):
        """Copy font and row heights from the source sheet to the Fixes sheet."""
        ws_source = self.wb[self.sourcesheet]  #####
        ws = self.wb[self.fixessheet]
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        for i in range(0, ws.max_row):
            cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
            for j in range(0, ws.max_column):
                col = cols[j]
                font_name = ws_source[f"{col}{i + 1}"].font.name
                font_size = ws_source[f"{col}{i + 1}"].font.size
                font_bold = ws_source[f"{col}{i + 1}"].font.bold
                font = Font(name=font_name, size=font_size, bold=font_bold)
                ws[f"{col}{i + 1}"].font = font
                ws[f"{col}{i + 1}"].alignment = wrap_alignment
            # [i + 1] - because the lines are numbered starting at 1
            ws.row_dimensions[i + 1].height = self._get_height_for_row(ws, i)

    def populate_summary(self):
        """Fill the Summary sheet with findings grouped by severity.

        Sections include Critical Errors, Please Check, Key Metrics, Possible
        Glitches, and Other Observations. Also sets ``self.summary``.

        Returns:
            dict: The same summary stored on ``self.summary``.
        """
        row = 1
        skip_rows = [1]
        ws = self.wb["Summary"]
        ws[f"A{row}"] = f"Config ID :{self.df_config_string}:"
        ws[f"A{row}"].style = "Explanatory Text"

        row += 1
        skip_rows.append(row)
        ws[f"A{row}"] = "Critical Errors"
        for c in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{c}{row}"].style = "Note"

        if len(self.diagnostics_table) > 0:
            df_diagnostics_table = pd.DataFrame(self.diagnostics_table)
        else:
            df_diagnostics_table = pd.DataFrame(
                {
                    "level": [],
                    "name": [],
                    "row": [],
                    "severity": [],
                    "cell": [],
                    "error": [],
                    "corrective_action": [],
                }
            )
        if len(self.diagnostics_column) > 0:
            df_diagnostics_column = pd.DataFrame(self.diagnostics_column)
        else:
            df_diagnostics_column = pd.DataFrame(
                {
                    "level": [],
                    "name": [],
                    "row": [],
                    "severity": [],
                    "cell": [],
                    "error": [],
                    "corrective_action": [],
                }
            )
        df = pd.concat(
            [
                df_diagnostics_table[df_diagnostics_table.severity >= 100],
                df_diagnostics_column[df_diagnostics_column.severity >= 100],
            ]
        ).sort_values(["severity", "row"], ascending=[False, True])[
            ["level", "name", "row", "severity", "cell", "error", "corrective_action"]
        ]

        if len(self.df) > 0:
            error_ratio = len(df) / len(self.df)
        else:
            error_ratio = 1
        self.summary = {
            "sourcefilename": self.sourcefilename,
            "targetfilename": self.targetfilename,
            "error_ratio": error_ratio,
            "num_lines": len(self.df),
            "num_errors": len(df),
        }
        row += 1
        if len(df) > 0:
            cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
            cnt = 0
            for c in list(df.columns) + ["Keep", "Value"]:
                column = cols[cnt]
                ws[f"{column}{row}"] = c
                cnt += 1

            row += 1

            for i, r in df.iterrows():
                cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
                cnt = 0
                for c in list(df.columns) + ["Keep", "Value"]:
                    column = cols[cnt]
                    if c == "Keep":
                        ws[f"{column}{row}"] = False
                    elif c == "Value":
                        if str(r["cell"]).lower() != "nan":
                            ws[f"{column}{row}"] = self.wb[self.sourcesheet][
                                r["cell"]
                            ].value
                    else:
                        if c == "cell" and str(r[c]).lower != "nan":
                            ws[
                                f"{column}{row}"
                            ].hyperlink = f"#'{self.fixessheet}'!{r[c]}"
                        ws[f"{column}{row}"] = r[c]
                    cnt += 1
                row += 1
        else:
            skip_rows.append(row)
            ws[f"A{row}"] = "no critical errors found"
            ws[f"A{row}"].style = "Good"

        row += 2
        skip_rows.append(row)
        ws[f"A{row}"] = f"Please Check"
        for c in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{c}{row}"].style = "Note"

        df = pd.concat(
            [
                df_diagnostics_table[
                    (df_diagnostics_table.severity >= 20)
                    & (df_diagnostics_table.severity < 100)
                ],
                df_diagnostics_column[
                    (df_diagnostics_column.severity >= 20)
                    & (df_diagnostics_column.severity < 100)
                ],
            ]
        ).sort_values(["severity", "row"], ascending=[False, True])[
            ["level", "name", "row", "severity", "cell", "error", "corrective_action"]
        ]

        row += 1
        if len(df) > 0:
            cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
            cnt = 0
            for c in list(df.columns) + ["Keep", "Value"]:
                column = cols[cnt]
                ws[f"{column}{row}"] = c
                cnt += 1
            row += 1

            for i, r in df.iterrows():
                cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
                cnt = 0
                for c in list(df.columns) + ["Keep", "Value"]:
                    column = cols[cnt]
                    if c == "Keep":
                        ws[f"{column}{row}"] = False
                    elif c == "Value":
                        if str(r["cell"]).lower() != "nan":
                            ws[f"{column}{row}"] = self.wb[self.sourcesheet][
                                r["cell"]
                            ].value
                    else:
                        if c == "cell" and str(r[c]).lower != "nan":
                            ws[
                                f"{column}{row}"
                            ].hyperlink = f"#'{self.fixessheet}'!{r[c]}"
                        ws[f"{column}{row}"] = r[c]
                    cnt += 1
                row += 1
        else:
            skip_rows.append(row)
            ws[f"A{row}"] = "Nothing to Check"
            ws[f"A{row}"].style = "Good"

        row += 2
        skip_rows.append(row)
        ws[f"A{row}"] = f"Key Metrics"
        for c in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{c}{row}"].style = "Note"

        df = pd.DataFrame(self.statistics)
        df = df[df.level == "table"]

        row += 1
        for i, r in df.iterrows():
            cnt = 0
            cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
            for c in ["level", "metric", "value"]:
                column = cols[cnt]
                ws[f"{column}{row}"] = r[c]
                cnt += 1
            row += 1
        df = pd.DataFrame(self.statistics)
        df = df[df.level == "column"]
        df = df[df.metric == "distinct_ratio_values"]
        df = df[df.name.isin(["country", "name", "address"])]
        first = FormatObject(type="min")
        mid = FormatObject(type="percentile", val=0.5)
        last = FormatObject(type="max")
        colors = [Color("F8696B"), Color("FFEB84"), Color("63BE7B")]
        cs3 = ColorScale(cfvo=[first, mid, last], color=colors)
        rule = Rule(type="colorScale", colorScale=cs3)
        rowfrom = row
        for i, r in df.iterrows():
            cnt = 0
            cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
            for c in ["level", "metric", "name", "value"]:
                column = cols[cnt]
                ws[f"{column}{row}"] = r[c]
                ws[f"{column}{row}"].style = "Percent"
                cnt += 1
            row += 1
        if row - 1 > rowfrom:
            ws.conditional_formatting.add(f"D{rowfrom}:D{row - 1}", rule)

        row += 2
        skip_rows.append(row)
        ws[f"A{row}"] = f"Possible Glitches"
        for c in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{c}{row}"].style = "Note"

        df = pd.concat(
            [
                df_diagnostics_table[
                    (df_diagnostics_table.severity > 1)
                    & (df_diagnostics_table.severity < 20)
                ],
                df_diagnostics_column[
                    (df_diagnostics_column.severity > 1)
                    & (df_diagnostics_column.severity < 20)
                ],
            ]
        ).sort_values(["severity", "row"], ascending=[False, True])[
            ["level", "name", "row", "severity", "cell", "error", "corrective_action"]
        ]

        row += 1
        if len(df) > 0:
            cnt = 0
            cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
            for c in list(df.columns) + ["Keep", "Value"]:
                column = cols[cnt]
                ws[f"{column}{row}"] = c
                cnt += 1

            row += 1
            for i, r in df.iterrows():
                cnt = 0
                cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
                for c in list(df.columns) + ["Keep", "Value"]:
                    column = cols[cnt]
                    if c == "Keep":
                        ws[f"{column}{row}"] = False
                    elif c == "Value":
                        if str(r["cell"]).lower() != "nan":
                            ws[f"{column}{row}"] = self.wb[self.sourcesheet][
                                r["cell"]
                            ].value
                    else:
                        if c == "cell" and str(r[c]).lower != "nan":
                            ws[
                                f"{column}{row}"
                            ].hyperlink = f"#'{self.fixessheet}'!{r[c]}"
                        ws[f"{column}{row}"] = r[c]
                    cnt += 1
                row += 1
        else:
            skip_rows.append(row)
            ws[f"A{row}"] = "Nothing looking like a glitch"
            ws[f"A{row}"].style = "Good"

        row += 2
        skip_rows.append(row)
        ws[f"A{row}"] = f"Other Observations"
        for c in ["A", "B", "C", "D", "E", "F", "G"]:
            ws[f"{c}{row}"].style = "Note"

        df = pd.concat(
            [
                df_diagnostics_table[(df_diagnostics_table.severity < 1)],
                df_diagnostics_column[(df_diagnostics_column.severity < 1)],
            ]
        ).sort_values(["severity", "row"], ascending=[False, True])[
            ["level", "name", "row", "severity", "cell", "error", "corrective_action"]
        ]

        row += 1
        cnt = 0
        cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
        for c in df.columns:
            column = cols[cnt]
            ws[f"{column}{row}"] = c
            cnt += 1

        row += 1
        for i, r in df.iterrows():
            cnt = 0
            cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
            for c in df.columns:
                column = cols[cnt]
                if c == "cell" and str(r[c]).lower() != "nan":
                    ws[f"{column}{row}"].hyperlink = f"#'{self.fixessheet}'!{r[c]}"
                ws[f"{column}{row}"] = r[c]
                cnt += 1
            row += 1

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 8.5
        ws.column_dimensions["E"].width = 6
        ws.column_dimensions["F"].width = 45
        ws.column_dimensions["G"].width = 45
        ws.column_dimensions["H"].width = 8
        ws.column_dimensions["I"].width = 45

        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        font = Font(size=10)

        for i in range(0, ws.max_row):
            if i + 1 in skip_rows:
                continue
            # [i + 1] - because the lines are numbered starting at 1
            cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
            for j in range(0, ws.max_column):
                col = cols[j]
                ws[f"{col}{i + 1}"].font = font
                ws[f"{col}{i + 1}"].alignment = wrap_alignment
            ws.row_dimensions[i + 1].height = self._get_height_for_row(
                ws, i, font_size=10
            )

        return self.summary

    def copy_values_and_layout(self, wb, df, source, target):
        """Copy data and cell styling from one sheet to another.

        Args:
            wb: The openpyxl workbook.
            df: DataFrame whose rows are written to the target sheet.
            source: Name of the sheet to copy layout from.
            target: Name of the sheet to populate.
        """
        self.widths = [
            wb[source].column_dimensions[c].width for c in self.mapping.values()
        ]
        self.heights = [wb[source].row_dimensions[1].height] + [
            wb[source].row_dimensions[r].height for r in list(df.index)
        ]

        self.populated_rows = []
        row_num = 1
        for r in dataframe_to_rows(df, index=False, header=True):
            wb[target].append(r)
            self.populated_rows.append(row_num)
            row_num += 1

        self.populated_columns = map_n_dataframe_cols_to_excel_cols(len(self.widths))
        for i in range(len(self.widths)):
            col = self.populated_columns[i]
            wb[target].column_dimensions[col].width = self.widths[i]

        wb[target].row_dimensions[1].height = self.heights[0]
        for i in range(len(df.index)):
            wb[target].row_dimensions[df.index[i]].height = self.heights[i]

        # This does not seem to work
        # https://openpyxl.readthedocs.io/en/latest/styles.html#applying-styles
        # Styles can also applied to columns and rows but note that this applies
        # only to cells created (in Excel) after the file is closed.
        for r in self.populated_rows:
            for c in self.populated_columns:
                print(f"values: {c} and {r}")
                if len(c) > 1:
                    continue
                alignment = copy(wb[source][f"{c}{r}"].alignment)
                wb[target][f"{c}{r}"].alignment = copy(wb[source][f"{c}{r}"].alignment)
                if alignment.wrap_text:
                    wb[target][f"{c}{r}"].alignment = Alignment(wrapText=True)

                font = copy(wb[source][f"{c}{r}"].font)
                wb[target][f"{c}{r}"].font = font

                border = copy(wb[source][f"{c}{r}"].border)
                wb[target][f"{c}{r}"].border = border

                number_format = copy(wb[source][f"{c}{r}"].number_format)
                wb[target][f"{c}{r}"].number_format = number_format

    def normalise_column_names(self, df):
        """Return a copy of ``df`` with lowercased column names."""
        rename = {}
        for column in df.columns:
            rename[column] = column.lower()
        return df.rename(columns=rename)

    def _add_stats(self, level="", name="", metric="", value=0):
        """Record a summary statistic for the Summary sheet."""
        self.statistics.append(
            {"level": level, "name": name, "metric": metric, "value": value}
        )

    def _add_metrics(self, level="", name="", metric="", value=0):
        """Record a column- or table-level metric during validation."""
        self.metrics.append(
            {"level": level, "name": name, "metric": metric, "value": value}
        )

    def _decide_autofix(self, code):
        """Return whether the config enables auto-fix for ``code``."""
        ddf = self.df_config[self.df_config.code == code]
        if len(ddf) > 0:
            return ddf.auto_fix.values[0]
        else:
            return False

    def _get_style_findings(self, code):
        """Return the Excel style name for highlighting a finding cell."""
        if self.df_config[self.df_config.code == code].style_findings.values[0] in [
            "Good",
            "Note",
            "Neutral",
            "Bad",
        ]:
            return self.df_config[self.df_config.code == code].style_findings.values[0]
        else:
            return ""

    def _get_style_fixed(self, code):
        """Return the Excel style name for a cell after auto-fix."""
        if self.df_config[self.df_config.code == code].style_fixed.values[0] in [
            "Good",
            "Note",
            "Neutral",
            "Bad",
        ]:
            return self.df_config[self.df_config.code == code].style_fixed.values[0]
        else:
            return ""

    def _get_highest_severity_style_fixed(self, codes):
        """Return the fixed-cell style for the highest-severity code in ``codes``."""
        ddf = self.df_config[self.df_config.code.isin(codes)].sort_values(
            "severity", ascending=False
        )
        if len(ddf) > 0:
            if ddf.style_fixed.values[0] in ["Good", "Note", "Neutral", "Bad"]:
                return ddf.style_fixed.values[0]
            else:
                return ""

    def _get_severity(self, code):
        """Return the configured severity for ``code``, or ``-1`` if unknown."""
        try:
            severity = self.df_config[self.df_config.code == code].severity.values[0]
        except:
            severity = -1
        return severity

    def _add_diagnosis(self, code="", column_name="", row="", **kwparams):
        """Record a table- or column-level finding and render its message.

        Message templates come from the config workbook and are rendered with
        Jinja2 using ``kwparams``. Table codes (``T*``) go to ``diagnostics_table``;
        column codes (``C*``) go to ``diagnostics_column``.

        Args:
            code: Error code (e.g. ``T0001``, ``C0001``).
            column_name: Column or context label for the finding.
            row: Excel row number or row pair for duplicate findings.
            **kwparams: Template variables for the error message.
        """
        env = NativeEnvironment()
        diagnostic_text = {"error": "", "corrective_action": "", "fixes_comment": ""}
        if code in self.df_config.code.values:
            ddf = self.df_config[self.df_config.code == code]
            severity = ddf.severity.values[0]
            for field in diagnostic_text.keys():
                if kwparams:
                    for k, v in kwparams.items():
                        kwparams[k] = str(v)
                    have_data = False
                    string = ddf[field].values[0]
                    if not isinstance(string, str):
                        string = ""
                    for entry in env.parse(string).body:
                        for node in entry.nodes:
                            if not isinstance(node, TemplateData):
                                if node.name in kwparams.keys():
                                    have_data = True
                                    break
                                else:
                                    want_data = node.name
                                    kwparams.update(
                                        {
                                            want_data: f"(#CfgErr: Cfg for code {code} has {want_data}, Code has {','.join(kwparams.keys())})"
                                        }
                                    )
                    if have_data:
                        result = env.from_string(string)
                    else:
                        result = env.from_string(string)
                else:
                    string = (
                        ddf[field].values[0]
                        if isinstance(ddf[field].values[0], str)
                        else ""
                    )
                    result = env.from_string(string)
                diagnostic_text[field] = result.render(kwparams)
        else:
            severity = 100
            diagnostic_text["error"] = (
                f"Tool error: No entry for field 'error' and error code {code}"
            )
            diagnostic_text["corrective_action"] = (
                "Provide a valid error_codes definition"
            )
            diagnostic_text["fixes_comment"] = ""

        if code.startswith("T"):
            self.diagnostics_table.append(
                {
                    "level": "table",
                    "code": code,
                    "name": column_name,
                    "row": row,
                    "severity": severity,
                    "error": diagnostic_text["error"],
                    "corrective_action": diagnostic_text["corrective_action"],
                    "fixes_comment": diagnostic_text["fixes_comment"],
                }
            )
        else:
            try:
                cell = f"{self.mapping[column_name]}{row}"
            except:
                cell = f"@{row}"
            self.diagnostics_column.append(
                {
                    "level": "column",
                    "code": code,
                    "name": column_name,
                    "row": row,
                    "cell": cell,
                    "severity": severity,
                    "error": diagnostic_text["error"],
                    "corrective_action": diagnostic_text["corrective_action"],
                    "fixes_comment": diagnostic_text["fixes_comment"],
                }
            )

        pass

    def _add_fix_comment(self, coordinates, code):
        """Queue a fix comment for a cell when auto-fix is enabled for ``code``."""
        if self._decide_autofix(code):
            self.fixes.append(
                {
                    "coordinates": coordinates,
                    "fix": self.df_config[
                        self.df_config.code == code
                    ].fixes_comment.values[0],
                }
            )

    def check_table(self):
        """Run table-level checks and collect column statistics.

        Validates required columns, lat/lng pairing, unrecognised columns, and
        empty tables. Populates ``diagnostics_table`` with ``T*`` codes.
        """
        self._add_stats(
            level="table", name=self.sourcesheet, metric="num_rows", value=len(self.df)
        )
        self._add_stats(
            level="table",
            name=self.sourcesheet,
            metric="num_columns",
            value=len(self.df.columns),
        )

        for c in [
            "country",
            "name",
            "address",
            "sector_product_type",
            "facility_type_processing_type",
            "number_of_workers",
            "parent_company",
            "lat",
            "lng",
        ]:
            if c not in self.df.columns:
                continue
            if len(self.df[c].dropna()) > 0:
                self._add_stats(
                    level="column",
                    name=c,
                    metric="num_values",
                    value=len(self.df[c].dropna()),
                )
            else:
                self._add_stats(
                    level="column", name=c, metric="num_values", value="empty"
                )

            if len(set(self.df[c].dropna().values)) > 0:
                self._add_stats(
                    level="column",
                    name=c,
                    metric="num_distinct_values",
                    value=len(set(self.df[c].dropna().values)),
                )
            else:
                self._add_stats(
                    level="column", name=c, metric="num_distinct_values", value="empty"
                )

            if (
                len(self.df[c].dropna()) > 0
                and len(set(self.df[c].dropna().values)) > 0
            ):
                self._add_stats(
                    level="column",
                    name=c,
                    metric="distinct_ratio_values",
                    value=len(set(self.df[c].dropna().values))
                    / len(self.df[c].dropna()),
                )
            else:
                self._add_stats(
                    level="column", name=c, metric="num_distinct_values", value="empty"
                )

        if "country" not in self.df.columns:
            self._add_diagnosis(code="T0001")

        if "name" not in self.df.columns:
            self._add_diagnosis(code="T0002")

        if "address" not in self.df.columns:
            self._add_diagnosis(code="T0003")

        if "sector_product_type" not in self.df.columns:
            self._add_diagnosis(code="T0004")

        if "lat" in self.df.columns and not "lng" in self.df.columns:
            self._add_diagnosis(code="T0005")
        elif "lng" in self.df.columns and not "lat" in self.df.columns:
            self._add_diagnosis(code="T0006")
        elif "lng" in self.df.columns and "lat" in self.df.columns:
            self._add_diagnosis(code="T0007")

        for column in self.df.columns:
            if column not in [
                "country",
                "name",
                "address",
                "sector_product_type",
                "facility_type_processing_type",
                "number_of_workers",
                "parent_company",
                "lat",
                "lng",
            ]:
                self._add_diagnosis(code="T0008", column=column)

        if len(self.df) < 1:
            self._add_diagnosis(code="T0009", column_name=self.sourcesheet)

    def _add_comments_to_excel_sheets(self):
        """Attach finding and fix comments to the Findings and Fixes sheets."""
        if len(self.diagnostics_column) > 0:
            df = pd.DataFrame(self.diagnostics_column).sort_values(
                ["severity", "name", "row"], ascending=[False, True, False]
            )
            for cell in df.cell.unique():
                ddf = df[df.cell == cell]

                findings_style = self._get_style_findings(ddf.code.values[0])
                if findings_style != "":
                    self.wb["Findings"][cell].style = findings_style

                fixes_style = self._get_style_fixed(ddf.code.values[0])
                if fixes_style != "":
                    self.wb[self.fixessheet][cell].style = fixes_style

                if len(ddf) > 1:
                    num_comments = f"{len(ddf)} Comments:\n"
                else:
                    num_comments = "1 Comment:\n"
                comment = Comment(
                    num_comments + "\n".join(ddf.error), "Data Quality Bot"
                )
                comment.height = 300
                comment.width = 400
                self.wb["Findings"][cell].comment = comment
                self.wb[self.fixessheet][cell].comment = comment

        for fix in self.fixes:
            comment_text = self.wb[self.fixessheet][fix["coordinates"]].comment.text
            if isinstance(comment, str):
                comment_text += "\nFix: " + fix["fix"]
            else:
                comment_text = "Fix: " + fix["fix"]
            comment = Comment(comment_text, "Data Quality Bot")
            comment.height = 300
            comment.width = 400
            self.wb[self.fixessheet][fix["coordinates"]].comment = comment

        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        font = Font(size=10)
        ws = self.wb["Findings"]
        for i in range(0, ws.max_row):
            # [i + 1] - because the lines are numbered starting at 1
            cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
            for j in range(0, ws.max_column):
                col = cols[j]
                ws[f"{col}{i + 1}"].font = font
                ws[f"{col}{i + 1}"].alignment = wrap_alignment
            ws.row_dimensions[i + 1].height = self._get_height_for_row(
                ws, i, font_size=10
            )

    def check_columns(self):
        """Run all column-level validators on the facility data.

        Delegates to individual ``check_column_*`` methods and duplicate detection.
        Normally invoked via :meth:`process` after :meth:`check_table`.
        """
        self.check_column_country()
        self.check_for_country_name_in_address()
        for column in [
            "country",
            "name",
            "address",
            "sector_product_type",
            "facility_type_processing_type",
            "number_of_workers",
            "parent_company",
            "lat",
            "lng",
        ]:
            if column not in self.df.columns:
                continue
            elif len(self.df[column]) == 0:
                self._add_diagnosis(
                    code="C0026", column_name=column, row=2, column=column
                )

            self.check_column_double_quotes(column)

        for column in [
            "country",
            "name",
            "address",
            "sector_product_type",
            "facility_type_processing_type",
            "number_of_workers",
            "parent_company",
            "lat",
            "lng",
        ]:
            if column not in self.df.columns:
                continue
            self.check_column_unusual_characters(column)

        for column in [
            "country",
            "name",
            "address",
            "sector_product_type",
            "facility_type_processing_type",
            "number_of_workers",
            "parent_company",
            "lat",
            "lng",
        ]:
            if column not in self.df.columns:
                continue
            self.check_column_leading_trailing_blanks(column)

        for column in [
            "country",
            "name",
            "address",
            "number_of_workers",
            "parent_company",
            "lat",
            "lng",
        ]:
            if column not in self.df.columns:
                continue
            self.check_column_lengths(column)

        for column in [
            "country",
            "name",
            "address",
            "sector_product_type",
            "facility_type_processing_type",
            "number_of_workers",
            "parent_company",
            "lat",
            "lng",
        ]:
            if column not in self.df.columns:
                continue
            self.check_column_excel_error_codes_na(column)

        for column in ["name", "address"]:
            if column not in self.df.columns:
                continue
            self.check_column_multiple_commas(column)

        for column in ["address"]:
            if column not in self.df.columns:
                continue
            self.check_column_po_box(column)

        for column in ["sector_product_type", "facility_type_processing_type"]:
            if column not in self.df.columns:
                continue
            self.check_column_correctly_delimited(column)

        self.check_name_address_duplicates()

    def check_column_country(self):
        """Validate country values against the known-countries list (``C0001``–``C0002``)."""
        if "country" in self.df.columns:
            row = 0
            # 2022-12-07 check for floats in column
            countries = []
            for c in self.df.country.values:
                if isinstance(c, str):
                    countries.append(c.lower().strip())
                else:
                    countries.append("")
            self._add_metrics(
                level="column",
                name="country",
                metric="number_different_values",
                value=len(set(countries)),
            )
            if len(countries) != 0:
                self._add_metrics(
                    level="column",
                    name="country",
                    metric="fraction_different_values",
                    value=len(set(countries)) / len(countries),
                )
            else:
                self._add_metrics(
                    level="column",
                    name="country",
                    metric="fraction_different_values",
                    value=0,
                )

            cells_with_errors = 0
            for row in range(len(countries)):
                if countries[row] not in self.known_countries:
                    self._add_diagnosis(
                        code="C0001",
                        column_name="country",
                        row=row + 2,
                        country=countries[row],
                    )
                    cells_with_errors += 1
                    row = self.df.index[row]
        else:
            cells_with_errors = len(self.df)
            self._add_diagnosis(code="C0002", column_name="country")

        if cells_with_errors > 0:
            self._add_metrics(
                level="column",
                name="country",
                metric="cells_with_errors",
                value=cells_with_errors,
            )

    def check_for_country_name_in_address(self):
        """Flag addresses whose trailing word looks like a mismatched country (``C0003``–``C0025``)."""
        us_states_of_concern = [
            "AL",
            "AZ",
            "AR",
            "CA",
            "CO",
            "DE",
            "GA",
            "ID",
            "IL",
            "IN",
            "KY",
            "LA",
            "ME",
            "MD",
            "MA",
            "MN",
            "MS",
            "MO",
            "MT",
            "NE",
            "NC",
            "PA",
            "SC",
            "SD",
            "TN",
            "VA",
        ]
        can_provinces_of_concern = ["PE", "SK", "NL"]
        if "country" in self.df.columns:
            cells_with_warnings = 0
            addresses = [
                a.lower() if isinstance(a, str) else "" for a in self.df.address
            ]
            countries = [
                c.lower().strip() if isinstance(c, str) else "" for c in self.df.country
            ]
            for row in range(len(addresses)):
                try:
                    last_word_in_address = re.findall(r".*\W(\w+)\W*$", addresses[row])[
                        0
                    ]
                except:
                    continue

                if last_word_in_address in self.known_countries:
                    country_code_address = COUNTRY_CODES[last_word_in_address]
                    try:
                        country_code_country = COUNTRY_CODES[countries[row]]
                    except:
                        continue  # fudged, this is for an unkinown country name
                    if country_code_address == country_code_country:
                        possible_problem = False
                    else:
                        if (
                            country_code_address in us_states_of_concern
                            and country_code_country == "US"
                        ):  # e.g. California/CA is also Canada(CA)
                            possible_problem = False
                        elif (
                            country_code_address in can_provinces_of_concern
                            and country_code_country == "CA"
                        ):  # e.g. California/CA is also Canada(CA)
                            possible_problem = False
                        else:
                            possible_problem = True
                    if (
                        len(re.findall("(?i).*(hong\s*kong)\s*$", addresses[row])) > 0
                        and COUNTRY_CODES[countries[row]] == "CN"
                    ):
                        self._add_diagnosis(
                            code="C0003", column_name="address", row=row + 2
                        )

                    elif (
                        len(re.findall("(?i).*(macao)\s*$", addresses[row])) > 0
                        and COUNTRY_CODES[countries[row]] == "MO"
                    ):
                        self._add_diagnosis(
                            code="C0004", column_name="address", row=row + 2
                        )

                    elif (
                        len(re.findall("(?i).*taiwan\s*$", addresses[row])) > 0
                        and COUNTRY_CODES[countries[row]] == "MO"
                    ):
                        self._add_diagnosis(
                            code="C0005", column_name="address", row=row + 2
                        )

                    elif possible_problem:
                        if len(last_word_in_address) > 3:
                            self._add_diagnosis(
                                code="C0006",
                                column_name="address",
                                row=row + 2,
                                country_in_address=last_word_in_address,
                                country=countries[row],
                            )
                        else:
                            self._add_diagnosis(
                                code="C0025",
                                column_name="address",
                                row=row + 2,
                                country_in_address=last_word_in_address,
                                country=countries[row],
                            )
                        cells_with_warnings += 1
        else:
            cells_with_warnings = 0

        if cells_with_warnings > 0:
            self._add_metrics(
                level="column",
                name="country",
                metric="cells_with_warnings",
                value=cells_with_warnings,
            )

    def check_column_unusual_characters(self, column):
        """Flag non-Latin or unusual Unicode characters in ``column`` (``C0007``).

        Args:
            column: Lowercased column name to inspect.
        """
        cells_with_warnings = 0

        values = self.df[column].values
        for row in range(len(values)):
            if not isinstance(values[row], str):
                continue
            unusual_characters = ""
            unusual_characters_names = set()
            for c in values[row]:
                if (
                    not unicodedata.category(c) == "Cc"
                    and not unicodedata.category(c) == "Co"
                    and not "LATIN" in unicodedata.name(c)
                    and not "SPACE" in unicodedata.name(c)
                    and not unicodedata.category(c) == "Po"
                    and not unicodedata.category(c) == "Sm"
                    and not unicodedata.category(c) == "Nd"
                    and not unicodedata.category(c) == "Pe"
                    and not unicodedata.category(c) == "Ps"
                    and not unicodedata.category(c) == "Pd"
                    and not unicodedata.category(c) == "Pf"
                    and not "MASCULINE ORDINAL INDICATOR" in unicodedata.name(c)
                    and not "ACCENT" in unicodedata.name(c)
                ):
                    unusual_characters += c
                    unusual_characters_names.add(unicodedata.name(c))
            if len(unusual_characters) > 0:
                unusual_characters = ",".join(unusual_characters)
                self._add_diagnosis(
                    code="C0007",
                    column_name=column,
                    row=row + 2,
                    unusual_characters=unusual_characters,
                    column=column,
                )
                cells_with_warnings += 1

        if cells_with_warnings > 0:
            self._add_metrics(
                level="column",
                name="country",
                metric="cells_with_warnings",
                value=cells_with_warnings,
            )

        return

    def check_column_double_quotes(self, column):
        """Detect and optionally remove double quotes in ``column`` (``C0008``).

        Args:
            column: Lowercased column name to inspect.
        """
        cells_with_errors = 0
        values = self.df[column].values
        for row in range(len(values)):
            if not isinstance(values[row], str):
                continue
            if '"' in values[row]:
                self._add_diagnosis(code="C0008", column_name=column, row=row + 2)
                cells_with_errors += 1
                coordinates = f"{self.mapping[column]}{row + 2}"

                if self._decide_autofix("C0008"):
                    self.wb[self.fixessheet][coordinates].value = (
                        self.wb[self.fixessheet][coordinates]
                        .value.replace('"', "")
                        .replace("  ", " ")
                    )
                    fixes_style = self._get_style_fixed("C0008")
                    if fixes_style != "":
                        self.wb[self.fixessheet][coordinates].style = fixes_style
                    self._add_fix_comment(coordinates=coordinates, code="C0008")

        if cells_with_errors > 0:
            self._add_metrics(
                level="column",
                name=column,
                metric="cells_with_errors",
                value=cells_with_errors,
            )

        return self.comments

    def check_column_po_box(self, column):
        """Detect post-office box patterns in ``column`` (``C0009``).

        Args:
            column: Lowercased column name to inspect (typically ``address``).
        """
        cells_with_errors = 0
        values = self.df[column].values
        for row in range(len(values)):
            if not isinstance(values[row], str):
                continue
            if re.match("(?i).*Post[ ]+Office.*", values[row]) or re.match(
                "(?i).*P\.*O\.*[\s]+Box.*", values[row]
            ):
                self._add_diagnosis(code="C0009", column_name=column, row=row + 2)
                cells_with_errors += 1
                coordinates = f"{self.mapping[column]}{row + 2}"

                if self._decide_autofix("C0009"):
                    self.wb[self.fixessheet][coordinates].value = self.wb[
                        self.fixessheet
                    ][coordinates].value.replace('"', "")
                    fixes_style = self._get_style_fixed("C0009")
                    if fixes_style != "":
                        self.wb[self.fixessheet][coordinates].style = fixes_style
                    self._add_fix_comment(coordinates=coordinates, code="C0009")

        if cells_with_errors > 0:
            self._add_metrics(
                level="column",
                name=column,
                metric="cells_with_errors",
                value=cells_with_errors,
            )

        return self.comments

    def _tokens_above_threshold(self, pair, threshold=1):
        """Return True when the numeric value in ``pair`` exceeds ``threshold``."""
        _, v = pair
        return v > threshold

    def check_column_leading_trailing_blanks(self, column):
        """Detect whitespace issues in ``column`` and apply auto-fixes (``C0010``–``C0015``).

        Checks for leading/trailing spaces, double spaces, tabs, carriage returns,
        and embedded newlines.

        Args:
            column: Lowercased column name to inspect.
        """
        cells_with_warnings = 0
        values = self.df[column].values
        codes = set()
        highest_severity_code_found = ""
        for row in range(len(values)):
            if not isinstance(values[row], str):
                continue
            found = False
            if values[row].startswith(" "):
                found = True
                self._add_diagnosis(code="C0010", column_name=column, row=row + 2)
                cells_with_warnings += 1
                coordinates = f"{self.mapping[column]}{row + 2}"
                codes.add("C0010")
                highest_severity_code_found = "C0010"

            if values[row].endswith(" "):
                found = True
                self._add_diagnosis(code="C0011", column_name=column, row=row + 2)
                cells_with_warnings += 1
                coordinates = f"{self.mapping[column]}{row + 2}"
                codes.add("C0011")
                if self._get_severity("C0011") > self._get_severity(
                    highest_severity_code_found
                ):
                    highest_severity_code_found = "C0011"

            if "  " in values[row]:
                found = True
                self._add_diagnosis(code="C0012", column_name=column, row=row + 2)
                cells_with_warnings += 1
                coordinates = f"{self.mapping[column]}{row + 2}"
                codes.add("C0012")
                if self._get_severity("C0012") > self._get_severity(
                    highest_severity_code_found
                ):
                    highest_severity_code_found = "C0012"

            if "\t" in values[row]:
                found = True
                self._add_diagnosis(code="C0013", column_name=column, row=row + 2)
                cells_with_warnings += 1
                coordinates = f"{self.mapping[column]}{row + 2}"
                codes.add("C0013")
                if self._get_severity("C0013") > self._get_severity(
                    highest_severity_code_found
                ):
                    highest_severity_code_found = "C0013"

            if "\r" in values[row]:
                found = True
                self._add_diagnosis(code="C0014", column_name=column, row=row + 2)
                cells_with_warnings += 1
                coordinates = f"{self.mapping[column]}{row + 2}"
                codes.add("C0014")
                if self._get_severity("C0014") > self._get_severity(
                    highest_severity_code_found
                ):
                    highest_severity_code_found = "C0014"

            if "\n" in values[row]:
                found = True
                self._add_diagnosis(code="C0015", column_name=column, row=row + 2)
                cells_with_warnings += 1
                coordinates = f"{self.mapping[column]}{row + 2}"
                codes.add("C0015")
                if self._get_severity("C0015") > self._get_severity(
                    highest_severity_code_found
                ):
                    highest_severity_code_found = "C0015"

            if found and (
                self._decide_autofix("C0010")
                or self._decide_autofix("C0011")
                or self._decide_autofix("C0012")
                or self._decide_autofix("C0013")
                or self._decide_autofix("C0014")
                or self._decide_autofix("C0015")
            ):
                new_value = self.wb[self.fixessheet][coordinates].value.strip()
                new_value = re.sub("[\n\r\t ,]+$", "", new_value)
                new_value = re.sub("^[\n\r\t ,]+", "", new_value)
                new_value = (
                    new_value.replace("\t", " ").replace("\r", " ").replace("\n", " ")
                )
                while "  " in new_value:
                    new_value = new_value.replace("  ", " ")
                self.wb[self.fixessheet][coordinates].value = new_value

                fixes_style = self._get_style_fixed(highest_severity_code_found)
                if fixes_style != "":
                    self.wb[self.fixessheet][coordinates].style = fixes_style
                for code in codes:
                    self._add_fix_comment(coordinates=coordinates, code=code)

        if cells_with_warnings > 0:
            self._add_metrics(
                level="column",
                name=column,
                metric="cells_with_warnings",
                value=cells_with_warnings,
            )

        return self.comments

    def check_column_lengths(self, column):
        """Flag values that are too long or too short for ``column`` (``C0016``–``C0018``).

        Args:
            column: Lowercased column name to inspect.
        """
        cells_with_errors = 0
        values = self.df[column].values
        for row in range(len(values)):
            if not isinstance(values[row], str):
                continue
            if len(values[row]) > 200:
                print("len(values[row])", len(values[row]))
                self._add_diagnosis(
                    code="C0016",
                    column_name=column,
                    row=row + 2,
                    cell_len=len(values[row]),
                )
                cells_with_errors += 1
            elif column == "name" and len(values[row]) < 10:
                self._add_diagnosis(
                    code="C0017",
                    column_name=column,
                    row=row + 2,
                    name_len=len(values[row]),
                )
                cells_with_errors += 1
            elif column == "address" and len(values[row]) < 25:
                self._add_diagnosis(
                    code="C0018",
                    column_name=column,
                    row=row + 2,
                    address_len=len(values[row]),
                )
                cells_with_errors += 1

        if cells_with_errors > 0:
            self._add_metrics(
                level="column",
                name=column,
                metric="cells_with_errors",
                value=cells_with_errors,
            )

        return self.comments

    def check_column_excel_error_codes_na(self, column):
        """Detect Excel error tokens such as ``#N/A`` in ``column`` (``C0019``).

        Args:
            column: Lowercased column name to inspect.
        """
        cells_with_warnings = 0
        values = self.df[column].values
        invalid_tokens = [
            "#DIV/0",
            "#N/A",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#REF!",
            "#VALUE!",
            "N/A",
            "#NA",
        ]
        for row in range(len(values)):
            if not isinstance(values[row], str):
                continue
            for invalid_token in invalid_tokens:
                if len(re.findall(r"\b{}\b".format(invalid_token), values[row])):
                    self._add_diagnosis(
                        code="C0019",
                        column_name=column,
                        row=row + 2,
                        invalid_token=invalid_token,
                    )
                    cells_with_warnings += 1
                    coordinates = f"{self.mapping[column]}{row + 2}"

                    self.wb[self.fixessheet][coordinates].value = self.wb[
                        self.fixessheet
                    ][coordinates].value.replace(invalid_token, "")
                    fixes_style = self._get_style_fixed("C0019")
                    if fixes_style != "":
                        self.wb[self.fixessheet][coordinates].style = fixes_style

        if cells_with_warnings > 0:
            self._add_metrics(
                level="column",
                name=column,
                metric="cells_with_warnings",
                value=cells_with_warnings,
            )

        return self.comments

    def check_column_multiple_commas(self, column):
        """Detect and normalise consecutive or empty comma-separated segments (``C0020``).

        Args:
            column: Lowercased column name to inspect (typically ``name`` or ``address``).
        """
        cells_with_warnings = 0
        values = self.df[column].values

        for row in range(len(values)):
            if not isinstance(values[row], str):
                continue
            num_consecutive_commas = len(re.findall(r",[ ]*,", values[row]))
            if num_consecutive_commas > 0:
                bits = []
                for bit in values[row].split(","):
                    if len(bit.strip()) > 0:
                        bits.append(bit.strip())
                new_value = ",".join(bits)
                self._add_diagnosis(
                    code="C0020",
                    column_name=column,
                    row=row + 2,
                    num_consecutive_commas=num_consecutive_commas,
                )
                cells_with_warnings += 1
                coordinates = f"{self.mapping[column]}{row + 2}"

                if self._decide_autofix("C0020"):
                    bits = []
                    for bit in self.wb[self.fixessheet][coordinates].value.split(","):
                        if len(bit.strip()) > 0:
                            bits.append(bit.strip())
                    new_value = ", ".join(bits)
                    self.wb[self.fixessheet][coordinates].value = new_value
                    fixes_style = self._get_style_fixed("C0020")
                    if fixes_style != "":
                        self.wb[self.fixessheet][coordinates].style = "Neutral"
                    self._add_fix_comment(coordinates=coordinates, code="C0020")

        if cells_with_warnings > 0:
            self._add_metrics(
                level="column",
                name=column,
                metric="cells_with_warnings",
                value=cells_with_warnings,
            )

        return self.comments

    def check_column_correctly_delimited(self, column):
        """Validate pipe-delimited multi-value fields in ``column`` (``C0021``–``C0024``).

        Ensures sector and facility-type columns use ``|`` rather than commas.

        Args:
            column: Lowercased column name to inspect.
        """
        cells_with_warnings = 0
        values = self.df[column].values

        for row in range(len(values)):
            if not isinstance(values[row], str):
                continue
            codes = set()
            if "|" in values[row] and "," in values[row]:
                bits = []
                for bit in values[row].split("|"):
                    if len(bit.strip()) > 0:
                        if "," in bit:
                            self._add_diagnosis(
                                code="C0021", column_name=column, row=row + 2
                            )
                            cells_with_warnings += 1
                            bit = re.sub(" +", " ", bit.replace(",", " "))
                            codes.add("C0021")
                        bits.append(bit.strip())
                coordinates = f"{self.mapping[column]}{row + 2}"
                new_value = "|".join(bits)
                self.wb[self.fixessheet][coordinates].value = new_value
                fixes_style = self._get_style_fixed("C0021")
                if fixes_style != "":
                    self.wb[self.fixessheet][coordinates].style = "Neutral"
            elif "," in values[row]:
                num_commas = values[row].count(",")
                num_fields = len(values[row].split("|"))
                if num_fields == 1:
                    self._add_diagnosis(
                        code="C0022",
                        column_name=column,
                        row=row + 2,
                        num_commas=num_commas,
                        column=column,
                    )
                    codes.add("C0022")
                else:
                    self._add_diagnosis(
                        code="C0023",
                        column_name=column,
                        row=row + 2,
                        num_commas=num_commas,
                        num_fields=num_fields,
                    )
                    codes.add("C0022")
                cells_with_warnings += 1
                coordinates = f"{self.mapping[column]}{row + 2}"

                new_value = self.wb[self.fixessheet][coordinates].value
                if num_commas < num_fields and self._decide_autofix(
                    "C0022"
                ):  # if commas seem intermixed
                    new_value = new_value.replace(",", "")
                    self._add_fix_comment(coordinates=coordinates, code="C0022")
                elif self._decide_autofix("C0023"):  # if only commas were used
                    new_value = new_value.replace(",", "|")
                    self._add_fix_comment(coordinates=coordinates, code="C0023")
                self.wb[self.fixessheet][coordinates].value = new_value

                style = self._get_highest_severity_style_fixed(list(codes))
                if style != "":
                    self.wb[self.fixessheet][coordinates].style = style

                bits = values[row].split("|")
            else:
                new_value = values[row]
                bits = values[row].split("|")

            if (
                len(new_value) - len(bits) - 1 > 200
            ):  # column lengths without | delimiters
                self._add_diagnosis(code="C0024", column_name=column, row=row + 2)
                cells_with_warnings += 1
                coordinates = f"{self.mapping[column]}{row + 2}"

        if cells_with_warnings > 0:
            self._add_metrics(
                level="column",
                name=column,
                metric="cells_with_warnings",
                value=cells_with_warnings,
            )

        return self.comments

    def _get_height_for_row(self, sheet, row_number, font_size=12):
        """Estimate row height from wrapped cell content and column width.

        Args:
            sheet: openpyxl worksheet.
            row_number: Zero-based row index.
            font_size: Font size used to pick height heuristics.

        Returns:
            float: Row height in Excel units.
        """
        factor_of_font_size_to_width = {
            # TODO: other sizes
            12: {
                "factor": 0.8,  # width / count of symbols at row
                "height": 16,
            },
            11: {
                "factor": 0.8,  # width / count of symbols at row
                "height": 14,
            },
            10: {
                "factor": 0.8,  # width / count of symbols at row
                "height": 13,
            },
        }
        try:
            font_params = factor_of_font_size_to_width[font_size]
        except:
            font_params = {"factor": 0.8, "height": 16}
        row = list(sheet.rows)[row_number]
        height = font_params["height"]

        for cell in row:
            words_count_at_one_row = (
                sheet.column_dimensions[cell.column_letter].width
                / font_params["factor"]
            )
            lines = ceil(len(str(cell.value)) / words_count_at_one_row)
            height = max(height, lines * font_params["height"])

        return height

    def check_name_address_duplicates(self):
        """Find similar or duplicate name/address pairs and populate Similarities (``T0010``–``T0014``).

        Compares all row pairs using fuzzy matching on normalised names and
        addresses. Writes results to the Similarities sheet and records table-level
        duplicate findings.
        """
        if len(self.df) <= 1:
            return

        addresses = []
        for a in self.df.address.values:
            if isinstance(a, str):
                addresses.append(a)
            else:
                addresses.append("")
        addresses_clean = [
            a.translate(str.maketrans("", "", string.punctuation)) for a in addresses
        ]
        addresses_lc = [a.lower() for a in addresses_clean]

        names = []
        for n in self.df.name.values:
            if isinstance(n, str):
                names.append(n)
            else:
                names.append("")
        names_clean = [
            n.translate(str.maketrans("", "", string.punctuation)) for n in names
        ]
        names_lc = [a.lower() for a in names_clean]
        names_cleaned = [cleanco.basename(n) for n in names_lc]
        names_cleaned = [cleanco.basename(n) for n in names_cleaned]

        df = self.df.join(
            pd.concat(
                [
                    self.df.name.str.extract(r"(?i)Unit[- ]*([0-9ixv]+)")
                    .dropna()
                    .join(self.df)[[0, "name"]]
                    .rename(columns={0: "unit_no"}),
                    self.df.name.str.extract(r"(?i)Factory[- ]*([0-9ixv]+)")
                    .dropna()
                    .join(self.df)[[0, "name"]]
                    .rename(columns={0: "unit_no"}),
                ]
            )[["unit_no"]]
        )
        df.unit_no.fillna(" ", inplace=True)

        alldata = []

        for i in tqdm(range(len(self.df))):
            for j in range(i + 1, len(self.df)):
                names_diff_contains_numbers = any(
                    c.isdigit()
                    for c in set(names_cleaned[i]).difference(names_cleaned[j])
                )
                addresses_ratio = (
                    rapidfuzz.fuzz.token_set_ratio(
                        unidecode(addresses_lc[i]), unidecode(addresses_lc[j])
                    )
                    / 100
                )
                names_ratio = (
                    rapidfuzz.fuzz.token_set_ratio(
                        unidecode(names_cleaned[i]), unidecode(names_cleaned[j])
                    )
                    / 100
                )
                if ((addresses_ratio > 0.8) | (names_ratio > 0.8)) & (
                    addresses_ratio * names_ratio > 0.75
                ):
                    alldata.append(
                        {
                            "i": i + 2,
                            "j": j + 2,
                            "address_1": addresses[i],
                            "address_2": addresses[j],
                            "name_1": names[i],
                            "name_2": names[j],
                            "unit_no_in_name_1": df.unit_no.values[i],
                            "unit_no_in_name_2": df.unit_no.values[j],
                            "addresses_ratio": addresses_ratio,
                            "names_ratio": names_ratio,
                            "ratio": addresses_ratio * names_ratio,
                            "names_diff_contains_numbers": names_diff_contains_numbers,
                        }
                    )
        if len(alldata) > 0:
            self.df_similarity = pd.DataFrame(alldata).sort_values(
                "ratio", ascending=False
            )
            self.df_similarity = self.df_similarity[
                (
                    (self.df_similarity.addresses_ratio > 0.8)
                    | (self.df_similarity.names_ratio > 0.8)
                )
                & (self.df_similarity.ratio > 0.75)
            ]

            df_duplicates = self.df_similarity[self.df_similarity.ratio >= 1.0]
            num_identical = len(df_duplicates)
        else:
            self.df_similarity = pd.DataFrame({"ratio": []})
            num_identical = 0
        #
        # Do not apply +2 offset, we already did that on computing the similarity
        if num_identical > 0:
            for i, row in df_duplicates.iterrows():
                self._add_diagnosis(
                    code="T0010",
                    column_name="duplicates",
                    row=f"{row.i} {row.j}",
                    row_1=row.i,
                    row_2=row.j,
                )

        df_duplicates = self.df_similarity[
            (self.df_similarity.ratio >= 0.9) & (self.df_similarity.ratio < 1.0)
        ]
        num_very_similar = len(df_duplicates)
        if num_very_similar > 0:
            for i, row in df_duplicates.iterrows():
                if (
                    len(row.unit_no_in_name_1) > 0 or len(row.unit_no_in_name_2)
                ) and row.unit_no_in_name_1 != row.unit_no_in_name_2:
                    self._add_diagnosis(
                        code="T0013", column_name="duplicates", row_1=row.i, row_2=row.j
                    )
                elif row.names_diff_contains_numbers:
                    self._add_diagnosis(
                        code="T0014", column_name="duplicates", row_1=row.i, row_2=row.j
                    )
                else:
                    self._add_diagnosis(
                        code="T0011", column_name="duplicates", row_1=row.i, row_2=row.j
                    )

        df_duplicates = self.df_similarity[(self.df_similarity.ratio < 0.9)]
        num_similar = len(df_duplicates)
        if num_similar > 0:
            for i, row in df_duplicates.iterrows():
                self._add_diagnosis(
                    code="T0012", column_name="duplicates", row_1=row.i, row_2=row.j
                )

        ws = self.wb["Similarities"]
        row = 1
        for r in dataframe_to_rows(self.df_similarity, index=False, header=True):
            ws.append(r)
            if row >= 2:
                ws[f"A{row}"].hyperlink = f"#'{self.sourcesheet}'!A{r[0]}"
                ws[f"B{row}"].hyperlink = f"#'{self.sourcesheet}'!A{r[1]}"
            row += 1

        for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            ws[f"{column}1"].style = "Headline 2"
            ws[f"{column}1"].alignment = Alignment(wrap_text=True, horizontal="center")

        row = 2
        for i, r in self.df_similarity.iterrows():
            for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
                ws[f"{column}{row}"].alignment = Alignment(
                    wrap_text=True, vertical="top"
                )
            if r.ratio >= 1.0:
                for column in [
                    "A",
                    "B",
                    "C",
                    "D",
                    "E",
                    "F",
                    "G",
                    "H",
                    "I",
                    "J",
                    "K",
                    "L",
                ]:
                    ws[f"{column}{row}"].style = "Bad"
                    ws[f"{column}{row}"].alignment = Alignment(
                        wrap_text=True, vertical="top"
                    )
            elif r.ratio >= 0.9:
                for column in [
                    "A",
                    "B",
                    "C",
                    "D",
                    "E",
                    "F",
                    "G",
                    "H",
                    "I",
                    "J",
                    "K",
                    "L",
                ]:
                    ws[f"{column}{row}"].style = "Neutral"
                    ws[f"{column}{row}"].alignment = Alignment(
                        wrap_text=True, vertical="top"
                    )
            row += 1

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 5
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 35
        ws.column_dimensions["F"].width = 35
        ws.column_dimensions["G"].width = 5
        ws.column_dimensions["H"].width = 5
        ws.column_dimensions["I"].width = 5
        ws.column_dimensions["J"].width = 5
        ws.column_dimensions["K"].width = 5
        ws.column_dimensions["L"].width = 10

        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        font = Font(size=10)

        for i in range(0, ws.max_row):
            # [i + 1] - because the lines are numbered starting at 1
            cols = map_n_dataframe_cols_to_excel_cols(ws.max_column)
            for j in range(0, ws.max_column):
                col = cols[j]
                ws[f"{col}{i + 1}"].font = font
                ws[f"{col}{i + 1}"].alignment = wrap_alignment
            ws.row_dimensions[i + 1].height = self._get_height_for_row(
                ws, i, font_size=10
            )

        self.num_identical = num_identical
        self.num_similar = num_similar
        self.num_lines = len(self.df)
        pass


if __name__ == "__main__":
    filename = sys.argv[-1]
    if "~PROCESSED" in filename or "~UPDATE" in filename:
        print(f"skipping processed file {filename}")
    elif os.path.exists(filename):
        bot = ContriBot(filename)
        bot.process()
        bot.save()
        pass
    else:
        print(f"cannot find file {filename}")
